# =====================================================
# ALBAT GRAPH V5 METIER COMPLETE
# PARTIE 1 / 3
# - structure complète
# - moteur universel
# - interface premium
# - espèces
# - contacts date complet
# =====================================================

import sys
import os
import json
import re
import math
import difflib
import calendar
import unicodedata
import pandas as pd
import subprocess

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    MATPLOTLIB_OK = True
except ImportError:
    MATPLOTLIB_OK = False

from PySide6.QtGui import QPixmap, QPalette, QBrush, Qt
from PySide6.QtCore import QTimer
from openpyxl.chart import (BarChart, LineChart, Reference, Series)
from openpyxl.chart.label import DataLabelList
from pathlib import Path
from openpyxl.utils import get_column_letter
from PySide6.QtWidgets import (
    QApplication, QWidget, QLabel,
    QPushButton, QFileDialog,
    QVBoxLayout, QMessageBox,
    QProgressBar, QComboBox, QHBoxLayout,
    QLineEdit, QFrame
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import (
    BarChart, LineChart, Reference
)
from openpyxl.chart.axis import ChartLines
from datetime import datetime, date, timedelta

from modules.utils import resource_path


ASSETS = resource_path("assets")
BG = resource_path("assets", "background_graph.png")
def open_file(path):
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.call(["open", path])
        else:
            subprocess.call(["xdg-open", path])
    except Exception:
        pass
def play_success_sound():
    try:
        sound_file = resource_path("success.wav")

        if sys.platform.startswith("win"):
            import winsound
            winsound.PlaySound(sound_file, winsound.SND_FILENAME)

        elif sys.platform == "darwin":
            subprocess.call(["afplay", sound_file])

        else:
            subprocess.call(["aplay", sound_file])

    except:
        pass


def autofit_worksheet(ws, min_width=10, max_width=45):
    for col_cells in ws.columns:
        max_len = 0
        col_index = col_cells[0].column
        col_letter = get_column_letter(col_index)

        for cell in col_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)
            except:
                pass

        width = max(min_width, min(max_len + 2, max_width))
        ws.column_dimensions[col_letter].width = width

def sun_time_fraction_to_hm(x):
    h = int(x)
    m = int((x - h) * 60)
    return h, m


def calc_sunset(dt, latitude, longitude):
    """
    Approximation NOAA coucher du soleil
    Retourne datetime local
    """

    n = dt.timetuple().tm_yday

    lng_hour = longitude / 15

    t = n + ((18 - lng_hour) / 24)

    M = (0.9856 * t) - 3.289

    L = M + (1.916 * math.sin(math.radians(M))) \
          + (0.020 * math.sin(math.radians(2 * M))) \
          + 282.634

    L = L % 360

    RA = math.degrees(math.atan(0.91764 * math.tan(math.radians(L))))
    RA = RA % 360

    Lquadrant  = (math.floor(L / 90)) * 90
    RAquadrant = (math.floor(RA / 90)) * 90
    RA = RA + (Lquadrant - RAquadrant)

    RA = RA / 15

    sinDec = 0.39782 * math.sin(math.radians(L))
    cosDec = math.cos(math.asin(sinDec))

    cosH = (
        math.cos(math.radians(90.833))
        - (sinDec * math.sin(math.radians(latitude)))
    ) / (
        cosDec * math.cos(math.radians(latitude))
    )

    if cosH > 1:
        return None

    H = math.degrees(math.acos(cosH))
    H = H / 15

    T = H + RA - (0.06571 * t) - 6.622

    UT = (T - lng_hour) % 24

    # France été/hiver simplifié
    offset = 2 if dt.month in [4,5,6,7,8,9,10] else 1

    local_hour = (UT + offset) % 24

    hh = int(local_hour)
    mm = int((local_hour - hh) * 60)

    return datetime(dt.year, dt.month, dt.day, hh, mm)
# =====================================================
# UNIVERSAL ENGINE
# =====================================================

ALIASES = {
    "species": [
        "espèce identifiée",
        "espece identifiee",
        "espèce",
        "espece",
        "species",
        "species name"
    ],
    "date": [
        "date",
        "day",
        "record date"
    ],
    "time": [
        "heure",
        "time",
        "hour"
    ],
    "temp": [
        "temp",
        "temperature",
        "température"
    ],
    "wind": [
        "vent",
        "wind",
        "wind speed"
    ]
}


def normalize(txt):

    txt = str(txt).lower().strip()

    txt = ''.join(
        c for c in unicodedata.normalize("NFD", txt)
        if unicodedata.category(c) != "Mn"
    )

    txt = txt.replace("_", " ")
    txt = txt.replace("-", " ")

    txt = re.sub(r"[^a-z0-9 ]", "", txt)
    txt = re.sub(r"\s+", " ", txt).strip()

    return txt


def find_column(df, key):

    vals = ALIASES.get(key, [key])

    best = None
    best_score = 0

    for col in df.columns:

        nc = normalize(col)

        for v in vals:

            nv = normalize(v)

            if nv == nc:
                return col

            score = difflib.SequenceMatcher(
                None, nv, nc
            ).ratio()

            if nv in nc:
                score = max(score, 0.95)

            if score > best_score:
                best_score = score
                best = col

    return best if best_score >= 0.65 else None


def clean_numeric(series):

    s = series.astype(str)
    s = s.str.replace(",", ".", regex=False)
    s = s.str.extract(
        r'([-+]?\d*\.?\d+)'
    )[0]

    return pd.to_numeric(
        s,
        errors="coerce"
    )


def smart_read_excel(path):

    xls = pd.ExcelFile(path)

    for sh in xls.sheet_names:

        for head in range(0, 10):

            try:
                df = pd.read_excel(
                    path,
                    sheet_name=sh,
                    header=head
                )

                if len(df.columns) < 2:
                    continue

                if find_column(
                    df,
                    "species"
                ):
                    return df

            except:
                pass

    return pd.read_excel(path)


# =====================================================
# UI
# =====================================================

class GraphWindow(QWidget):
    def paintEvent(self, event):
        from PySide6.QtGui import QPainter

        painter = QPainter(self)

        if self.bg_pix and not self.bg_pix.isNull():
            scaled = self.bg_pix.scaled(
                self.size(),
                Qt.KeepAspectRatioByExpanding,
                Qt.SmoothTransformation
            )

            x = (self.width() - scaled.width()) // 2
            y = (self.height() - scaled.height()) // 2

            painter.drawPixmap(x, y, scaled)
    def __init__(self, park_name="", ref_city="", ref_year=None, bureau_etude="", default_save_dir="", correlations_dir_hint=""):

        super().__init__()
        
        self.park_name = park_name
        self.ref_city = ref_city
        self.ref_year = ref_year
        self.bureau_etude = bureau_etude
        self.default_save_dir = default_save_dir
        self.correlations_dir = correlations_dir_hint
        self.file = ""
        self.city_name_found = None
        self.city_lat = None
        self.city_lon = None

        self.setWindowTitle(
            "Albat Graph"
        )
        self.bg_pix = QPixmap(BG) if os.path.exists(BG) else None   
        self.resize(500, 900)
        self.setMinimumSize(500, 900)
             
        self.setStyleSheet("""
        QWidget{
            
            background:transparent;
            color:white;
            font-family:"Segoe UI","Helvetica Neue",Arial,sans-serif;
        }
        QComboBox{
            background:black;
            color:#eef7d0;
            border-radius:12px;
            padding:8px;
            border:1px solid rgba(255,255,255,45);
        }

        QComboBox QAbstractItemView{
            background:black;
            color:#eef7d0;
            selection-background-color:#5a7f47;
            selection-color:white;
            border:1px solid rgba(255,255,255,35);
        }                   
        QPushButton{
            background:rgba(90,127,71,145);
            border-radius:14px;
            padding:14px;
            font-size:18px;
            font-weight:700;
        }

        QPushButton:hover{
            background:#6b9654;
        }

        QLabel{
            font-size:16px;
        }
                           
        QProgressBar{
            background:black;
            border:1px solid rgba(255,255,255,35);
            border-radius:12px;
            height:24px;
            text-align:center;
            color:#eef7d0;
        }          
       QProgressBar::chunk{
            background:#a6f56f;
            border-radius:11px;
        }
        """)

        lay = QVBoxLayout(self)
        lay.setSpacing(16)
        lay.setContentsMargins(
            35, 40, 35, 40
        )
# ===============================
# AJOUT CHOIX VILLE
# ===============================

        # Le grand titre décoratif "Albat / GRAPH" a été retiré :
        # l'onglet actif dans la barre d'onglets du haut suffit
        # déjà à indiquer où l'on se trouve, pas besoin de le
        # répéter en grand dans le contenu de chaque onglet.
        lay.addSpacing(10)

        zone = QVBoxLayout()

        lab_city = QLabel("Ville de référence (lever / coucher du soleil) :")

        city_line = QHBoxLayout()

        self.city_input = QLineEdit()
        self.city_input.setPlaceholderText("Ville de référence (ex : Dijon)")
        self.city_input.returnPressed.connect(self.search_city)

        btn_city = QPushButton("Rechercher")
        btn_city.setStyleSheet("""
            QPushButton{
                background:rgba(120,150,90,180);
                border:1px solid rgba(255,255,255,35);
                border-radius:12px;
                color:white;
                font-size:13px;
                font-weight:700;
                padding:8px 16px;
            }
            QPushButton:hover{
                background:rgba(160,210,120,220);
            }
            QPushButton:pressed{
                background:rgba(90,120,70,255);
            }
        """)
        btn_city.clicked.connect(self.search_city)

        city_line.addWidget(self.city_input)
        city_line.addWidget(btn_city)

        self.city_status = QLabel("Aucune ville recherchée.")
        self.city_status.setWordWrap(True)
        self.city_status.setStyleSheet("""
            color: #eef7d0;
            font-size: 11px;
            background:transparent;
        """)

        zone.addWidget(lab_city)
        zone.addLayout(city_line)
        zone.addWidget(self.city_status)

        
        
        lay.addStretch()
        lay.addLayout(zone)
        lay.addSpacing(10)
        

        self.lbl = QLabel(
            "Aucun fichier sélectionné"
        )

        self.pb = QProgressBar()

        b1 = QPushButton("")
        b1.setFixedSize(320, 90)

        zone_btn = QVBoxLayout(b1)
        zone_btn.setContentsMargins(0, 10, 0, 10)
        zone_btn.setSpacing(2)

        txt1 = QLabel("Choisir fichier Excel")
        txt1.setAlignment(Qt.AlignCenter)
        txt1.setAttribute(Qt.WA_TransparentForMouseEvents)
        txt1.setStyleSheet("""
        font-size:20px;
        font-weight:800;
        color:#eef7d0;
        background:transparent;
        """)

        txt2 = QLabel("Fichiers excel déjà corrélés acoustique/météo")
        txt2.setAlignment(Qt.AlignCenter)
        txt2.setAttribute(Qt.WA_TransparentForMouseEvents)
        txt2.setStyleSheet("""
        font-size:10px;
        color:rgba(255,255,255,180);
        background:transparent;
        """)

        zone_btn.addWidget(txt1)
        zone_btn.addWidget(txt2)

        b1.clicked.connect(self.pick)

        run_btn_style = """
            QPushButton{
                background:rgba(90,127,71,145);
                border:1px solid rgba(255,255,255,40);
                border-radius:16px;
                padding:14px;
                font-size:15px;
                font-weight:700;
                color:white;
            }
            QPushButton:hover{
                background:rgba(130,184,99,220);
            }
        """

        btn_width = 390

        b2 = QPushButton("GENERER RAPPORT EXCEL")
        b2.setMinimumHeight(52)
        b2.setFixedWidth(btn_width)
        b2.setStyleSheet(run_btn_style)
        b2.clicked.connect(lambda: self.run())

        b3 = QPushButton("EXPORTER LES GRAPHIQUES EN IMAGES")
        b3.setMinimumHeight(52)
        b3.setFixedWidth(btn_width)
        b3.setStyleSheet(run_btn_style)
        b3.clicked.connect(lambda: self.export_graphs_images())

        b4 = QPushButton("EXPORTER LES TABLEAUX EN IMAGES")
        b4.setMinimumHeight(52)
        b4.setFixedWidth(btn_width)
        b4.setStyleSheet(run_btn_style)
        b4.clicked.connect(lambda: self.export_tables_images())

        lay.addWidget(b1, alignment=Qt.AlignCenter)
        lay.addWidget(self.lbl)
        lay.addWidget(b2, alignment=Qt.AlignCenter)
        lay.addWidget(b3, alignment=Qt.AlignCenter)
        lay.addWidget(b4, alignment=Qt.AlignCenter)

        # --- Séparateur visuel avant le bouton "Tout générer" ---
        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setFixedWidth(btn_width)
        separator.setStyleSheet(
            "background:rgba(255,255,255,50); "
            "border:none; max-height:1px; min-height:1px;"
        )

        sep_wrap = QVBoxLayout()
        sep_wrap.setContentsMargins(0, 14, 0, 14)
        sep_wrap.addWidget(separator, alignment=Qt.AlignCenter)
        lay.addLayout(sep_wrap)

        b5 = QPushButton("🗂️ TOUT GÉNÉRER")
        b5.setMinimumHeight(52)
        b5.setFixedWidth(btn_width)
        b5.setStyleSheet("""
            QPushButton{
                background:rgba(184,141,58,150);
                border:1px solid rgba(255,214,140,90);
                border-radius:16px;
                padding:14px;
                font-size:15px;
                font-weight:700;
                color:white;
            }
            QPushButton:hover{
                background:rgba(214,166,79,220);
            }
        """)
        b5.clicked.connect(self.generate_all)

        lay.addWidget(b5, alignment=Qt.AlignCenter)

        lay.addWidget(self.pb)

        # Pré-remplissage ET validation automatique de la ville de
        # référence transmise depuis le popup de démarrage du
        # projet : évite d'avoir à recliquer sur "Rechercher" à
        # chaque ouverture d'onglet. L'appel réseau reste protégé
        # (try/except dans search_city -> _lookup_city_geoapi) : en
        # cas d'échec (pas de connexion), l'onglet reste utilisable,
        # il faudra juste valider la ville manuellement.
        if self.ref_city:
            self.city_input.setText(self.ref_city)
            self.search_city()

        # Auto-chargement du fichier Excel du dossier "Correlations"
        # du projet, pour ne pas avoir à le sélectionner à nouveau
        # manuellement dans chaque onglet (Graph, Bridage, Scenar).
        # Le premier fichier Excel valide trouvé est utilisé ; en
        # cas d'ambiguïté (plusieurs fichiers), l'utilisateur garde
        # la main via le bouton de sélection habituel.
        auto_file = self._find_correlations_excel()
        if auto_file:
            self.file = auto_file
            self.lbl.setText(os.path.basename(auto_file))

    # ======================================================

    def _find_correlations_excel(self):
        """
        Cherche, dans le dossier 'Correlations' du projet, un
        fichier Excel (.xlsx/.xlsm/.xls) à charger automatiquement.
        Ignore les fichiers verrous (~$...) et cachés. Retourne le
        premier trouvé (ordre alphabétique), ou None si absent.
        """

        if not self.correlations_dir or not os.path.isdir(
            self.correlations_dir
        ):
            return None

        try:
            fichiers = sorted(os.listdir(self.correlations_dir))
        except Exception:
            return None

        for f in fichiers:
            if f.startswith("~$") or f.startswith("."):
                continue
            if f.lower().endswith((".xlsx", ".xlsm", ".xls")):
                return os.path.join(self.correlations_dir, f)

        return None


    def _lookup_city_geoapi(self, name):
        """
        Recherche une commune française via l'API officielle
        geo.api.gouv.fr. Couvre les ~35000 communes de France,
        contrairement à une liste figée de grandes villes.

        Retourne un tuple (nom, latitude, longitude, departement)
        ou None si la ville est introuvable / pas de connexion.
        """

        import urllib.request
        import urllib.parse
        import json as _json

        try:
            query = urllib.parse.urlencode({
                "nom": name,
                "fields": "nom,centre,population,codeDepartement",
                "boost": "population",
                "limit": 1,
            })
            url = f"https://geo.api.gouv.fr/communes?{query}"

            req = urllib.request.Request(
                url, headers={"User-Agent": "AlbatGraph/1.0"}
            )

            with urllib.request.urlopen(req, timeout=6) as response:
                data = _json.loads(response.read().decode("utf-8"))

            if not data:
                return None

            commune = data[0]
            lon, lat = commune["centre"]["coordinates"]

            return (
                commune["nom"],
                lat,
                lon,
                commune.get("codeDepartement", "?")
            )

        except Exception:
            return None

    def search_city(self):

        name = self.city_input.text().strip()

        if not name:
            self.city_status.setText("Veuillez saisir un nom de ville.")
            return

        self.city_status.setText("Recherche en ligne...")
        QApplication.processEvents()

        result = self._lookup_city_geoapi(name)

        if result is not None:

            found_name, lat, lon, dept = result

            self.city_name_found = found_name
            self.city_lat = lat
            self.city_lon = lon

            self.city_status.setText(
                f"Trouvée : {found_name} / France (dép. {dept}) "
                f"(lat {lat:.2f}, lon {lon:.2f})"
            )

        else:

            self.city_name_found = None
            self.city_lat = None
            self.city_lon = None

            self.city_status.setText(
                "Ville introuvable. Vérifiez l'orthographe, ou votre "
                "connexion internet (la recherche interroge la base "
                "officielle des communes de France)."
            )

    def pick(self):

        f, _ = QFileDialog.getOpenFileName(
            self,
            "Choisir fichier",
            "",
            "Excel (*.xlsx *.xlsm *.xls)"
        )

        if f:
            self.file = f
            self.lbl.setText(
                os.path.basename(f)
            )


    def export_graphs_images(self, target_dir=None):
        """
        Régénère les mêmes graphiques que le rapport Excel, mais
        avec matplotlib (fonctionne identiquement sur Windows et
        Mac), et les enregistre en .jpg dans un dossier choisi (ou
        directement dans target_dir si fourni, sans dialogue —
        utilisé par 'Tout générer').
        """

        try:

            if not MATPLOTLIB_OK:
                raise Exception(
                    "Le module 'matplotlib' n'est pas installé.\n"
                    "Ouvrez un terminal et exécutez :\n"
                    "pip install matplotlib\n\n"
                    "puis relancez l'application."
                )

            if self.city_lat is None or self.city_lon is None:
                raise Exception(
                    "Recherchez et validez une ville de référence "
                    "avant d'exporter les graphiques."
                )

            if not self.file:
                raise Exception("Sélectionnez un fichier.")

            lat, lon = self.city_lat, self.city_lon

            if target_dir:
                out_dir = target_dir
                if not os.path.isdir(out_dir):
                    raise Exception(
                        "Dossier de destination introuvable :\n"
                        f"{out_dir}"
                    )
            else:
                out_dir = QFileDialog.getExistingDirectory(
                    self, "Choisir le dossier de sortie pour les images",
                    self.default_save_dir or ""
                )

                if not out_dir:
                    return

            self.pb.setValue(10)

            df = smart_read_excel(self.file)

            species_col = find_column(df, "species")
            date_col = find_column(df, "date")
            time_col = find_column(df, "time")
            temp_col = find_column(df, "temp")
            wind_col = find_column(df, "wind")

            if not species_col:
                raise Exception("Colonne espèce introuvable.")

            if date_col:
                df[date_col] = pd.to_datetime(df[date_col], errors="coerce")

            year = 2025

            if date_col and not df[date_col].dropna().empty:
                year = int(df[date_col].dt.year.mode().iloc[0])

            if self.ref_year:
                year = self.ref_year

            park_part = self.park_name if self.park_name else "Albat"
            safe_park = re.sub(r'[\\/*?:"<>|]', "_", park_part)

            if not self._confirm_overwrite_batch(
                out_dir, f"{safe_park}_{year}_"
            ):
                return

            plt.rcParams.update({
                "figure.facecolor": "white",
                "axes.facecolor": "white",
                "font.size": 10,
            })

            saved = []
            n = 0

            def save_fig(fig, name):
                nonlocal n
                n += 1
                path = os.path.join(
                    out_dir, f"{safe_park}_{year}_{n:02d}_{name}.jpg"
                )
                fig.savefig(path, dpi=150, bbox_inches="tight")
                plt.close(fig)
                saved.append(path)

            # ---- 1. Contacts par espèce ----
            tab = (
                df.groupby(species_col).size()
                .reset_index(name="Contacts")
                .sort_values("Contacts", ascending=False)
            )
            fig, ax = plt.subplots(figsize=(9, 5))
            palette1 = plt.cm.tab20.colors
            bar_colors1 = [palette1[i % len(palette1)] for i in range(len(tab))]
            ax.bar(tab[species_col].astype(str), tab["Contacts"], color=bar_colors1)
            ax.set_ylabel("Nombre de contacts")
            ax.tick_params(axis="x", rotation=45)
            fig.tight_layout()
            save_fig(fig, "Contacts_par_espece")

            self.pb.setValue(25)

            # ---- 2. Contacts par date ----
            if date_col:
                daily = (
                    df.dropna(subset=[date_col])
                    .groupby(df[date_col].dt.date).size()
                    .reset_index(name="Contacts")
                )
                fig, ax = plt.subplots(figsize=(10, 4))
                ax.plot(daily[date_col], daily["Contacts"], color="#2E86AB", linewidth=1.8)
                ax.set_ylabel("Nombre de contacts")
                fig.autofmt_xdate()
                fig.tight_layout()
                save_fig(fig, "Contacts_par_date")

                # ---- 3. Contacts par mois ----
                mois_fr = {
                    1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril",
                    5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août",
                    9: "Septembre", 10: "Octobre", 11: "Novembre",
                    12: "Décembre"
                }
                monthly = (
                    df.dropna(subset=[date_col])
                    .groupby(df[date_col].dt.month).size()
                    .reset_index(name="Contacts")
                )
                labels = [
                    mois_fr.get(int(m), str(m))
                    for m in monthly[date_col]
                ]
                fig, ax = plt.subplots(figsize=(8, 5))
                ax.bar(labels, monthly["Contacts"], color="#E67E22")
                ax.set_ylabel("Nombre de contacts")
                ax.tick_params(axis="x", rotation=45)
                fig.tight_layout()
                save_fig(fig, "Contacts_par_mois")

            self.pb.setValue(45)

            # ---- 4. Contacts par espèce et mois ----
            if date_col:
                tmp = df.dropna(subset=[date_col]).copy()
                tmp["Mois"] = tmp[date_col].dt.month
                pivot = tmp.pivot_table(
                    index=species_col, columns="Mois",
                    aggfunc="size", fill_value=0
                )
                fig, ax = plt.subplots(figsize=(11, 6))
                pivot.T.plot(kind="bar", stacked=True, ax=ax, colormap="tab20")
                ax.set_xlabel("")
                ax.set_ylabel("Nombre de contacts")
                ax.legend(loc="upper left", bbox_to_anchor=(1, 1), fontsize=8)
                fig.tight_layout()
                save_fig(fig, "Contacts_par_espece_et_mois")

                # ---- 5. Activité Noctule/Sérotine/Pipistrelles ----
                mois_range = range(4, 12)
                g1_counts, g2_counts, labels5 = [], [], []
                for mois in mois_range:
                    sub = tmp[tmp["Mois"] == mois]
                    g1 = sub[sub[species_col].str.contains(
                        "Noctule|Sérotine|Serotine", case=False, na=False
                    )]
                    g2 = sub[sub[species_col].str.contains(
                        "Pipistrelle|Pipistrelles", case=False, na=False
                    )]
                    g1_counts.append(len(g1))
                    g2_counts.append(len(g2))
                    labels5.append(mois_fr.get(mois, str(mois)))

                x = range(len(labels5))
                fig, ax = plt.subplots(figsize=(9, 5))
                ax.bar(
                    [i - 0.2 for i in x], g1_counts, width=0.4,
                    label="Noctule + Sérotine", color="#2E86AB"
                )
                ax.bar(
                    [i + 0.2 for i in x], g2_counts, width=0.4,
                    label="Pipistrelles", color="#E67E22"
                )
                ax.set_xticks(list(x))
                ax.set_xticklabels(labels5, rotation=45)
                ax.set_ylabel("Nombre de contacts")
                ax.legend()
                fig.tight_layout()
                save_fig(fig, "Activite_taxons")

            self.pb.setValue(65)

            # ---- 6. Contacts selon vent ----
            if wind_col:
                v = clean_numeric(df[wind_col]).dropna()
                if not v.empty:
                    bins = [x / 2 for x in range(0, 25)]
                    grp = pd.cut(v, bins=bins, right=False).value_counts().sort_index()
                    labels6 = [f"{it.left:g}-{it.right:g}" for it in grp.index]
                    fig, ax = plt.subplots(figsize=(9, 5))
                    ax.bar(labels6, grp.values, color="#17A2A0")
                    ax.set_xlabel("Vitesse de vent (m/s)")
                    ax.set_ylabel("Nombre de contacts")
                    ax.tick_params(axis="x", rotation=90)
                    fig.tight_layout()
                    save_fig(fig, "Contacts_selon_vent")

            # ---- 7. Contacts selon température ----
            if temp_col:
                t = clean_numeric(df[temp_col]).dropna()
                if not t.empty:
                    bins = list(range(int(t.min()), int(t.max()) + 2))
                    grp = pd.cut(t, bins=bins, right=False).value_counts().sort_index()
                    labels7 = [f"{int(it.left)}-{int(it.right)}" for it in grp.index]
                    fig, ax = plt.subplots(figsize=(9, 5))
                    ax.bar(labels7, grp.values, color="#D9534F")
                    ax.set_xlabel("Température (°C)")
                    ax.set_ylabel("Nombre de contacts")
                    ax.tick_params(axis="x", rotation=90)
                    fig.tight_layout()
                    save_fig(fig, "Contacts_selon_temperature")

            self.pb.setValue(85)

            # ---- 8. Contacts après coucher du soleil ----
            if date_col and time_col:
                tmp2 = df.copy()
                tmp2[date_col] = pd.to_datetime(tmp2[date_col], errors="coerce")
                tmp2[time_col] = pd.to_datetime(
                    tmp2[time_col].astype(str), errors="coerce"
                )
                compte = {i: 0 for i in range(13)}
                for _, row in tmp2.iterrows():
                    if pd.isna(row[date_col]) or pd.isna(row[time_col]):
                        continue
                    d = row[date_col]
                    heure_contact = datetime(
                        d.year, d.month, d.day,
                        row[time_col].hour, row[time_col].minute,
                        row[time_col].second
                    )
                    sunset = calc_sunset(d, lat, lon)
                    if sunset is None:
                        continue
                    if heure_contact.hour < 12:
                        heure_contact += timedelta(days=1)
                    heures = (heure_contact - sunset).total_seconds() / 3600
                    if heures < 0 or heures > 12:
                        continue
                    classe = int(heures)
                    if classe in compte:
                        compte[classe] += 1

                fig, ax = plt.subplots(figsize=(9, 5))
                ax.bar(
                    [f"H+{h}" for h in range(13)],
                    [compte[h] for h in range(13)],
                    color="#6C5CE7"
                )
                ax.set_xlabel("Heures après le coucher du soleil")
                ax.set_ylabel("Nombre de contacts")
                fig.tight_layout()
                save_fig(fig, "Contacts_apres_coucher")

            self.pb.setValue(100)

            QMessageBox.information(
                self, "Succès",
                f"{len(saved)} image(s) enregistrée(s) dans :\n{out_dir}"
            )

        except Exception as e:
            QMessageBox.critical(
                self, "Erreur",
                f"Impossible d'exporter les graphiques :\n{e}"
            )

    # ======================================================

    def export_tables_images(self, target_dir=None):
        """
        Régénère les mêmes tableaux que le rapport Excel, mais
        rendus en image (matplotlib, identique sur Windows/Mac),
        et les enregistre en .jpg dans un dossier choisi (ou
        directement dans target_dir si fourni, sans dialogue —
        utilisé par 'Tout générer').
        """

        try:

            if not MATPLOTLIB_OK:
                raise Exception(
                    "Le module 'matplotlib' n'est pas installé.\n"
                    "Ouvrez un terminal et exécutez :\n"
                    "pip install matplotlib\n\n"
                    "puis relancez l'application."
                )

            if self.city_lat is None or self.city_lon is None:
                raise Exception(
                    "Recherchez et validez une ville de référence "
                    "avant d'exporter les tableaux."
                )

            if not self.file:
                raise Exception("Sélectionnez un fichier.")

            if target_dir:
                out_dir = target_dir
                if not os.path.isdir(out_dir):
                    raise Exception(
                        "Dossier de destination introuvable :\n"
                        f"{out_dir}"
                    )
            else:
                out_dir = QFileDialog.getExistingDirectory(
                    self, "Choisir le dossier de sortie pour les images",
                    self.default_save_dir or ""
                )

                if not out_dir:
                    return

            self.pb.setValue(10)

            df = smart_read_excel(self.file)

            species_col = find_column(df, "species")
            date_col = find_column(df, "date")
            temp_col = find_column(df, "temp")
            wind_col = find_column(df, "wind")

            if not species_col:
                raise Exception("Colonne espèce introuvable.")

            if date_col:
                df[date_col] = pd.to_datetime(df[date_col], errors="coerce")

            year = 2025

            if date_col and not df[date_col].dropna().empty:
                year = int(df[date_col].dt.year.mode().iloc[0])

            if self.ref_year:
                year = self.ref_year

            park_part = self.park_name if self.park_name else "Albat"
            safe_park = re.sub(r'[\\/*?:"<>|]', "_", park_part)

            if not self._confirm_overwrite_batch(
                out_dir, f"{safe_park}_{year}_T"
            ):
                return

            mois_fr = {
                1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril",
                5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août",
                9: "Septembre", 10: "Octobre", 11: "Novembre",
                12: "Décembre"
            }

            saved = []
            n = 0

            def save_table(headers, rows, title, name, col_widths=None):
                nonlocal n
                n += 1

                ncols = len(headers)
                nrows = len(rows)

                # Largeur de chaque colonne proportionnelle à la
                # longueur du texte le plus long qu'elle contient
                # (en-tête compris) : sans ça, matplotlib partage
                # la largeur à parts égales entre colonnes, ce qui
                # tronque les colonnes à texte long (ex :
                # "Indicateur", "Espèce") au profit de colonnes de
                # chiffres bien plus courtes.
                max_lens = [
                    max(
                        [len(str(h))]
                        + [len(str(row[j])) for row in rows]
                    )
                    for j, h in enumerate(headers)
                ]

                if col_widths is None:
                    total_len = sum(max_lens) or 1
                    # Plancher à 6% pour qu'une colonne courte (ex :
                    # un mois "Juil") reste lisible même si une
                    # autre colonne du même tableau est très longue.
                    col_widths = [
                        max(length / total_len, 0.06)
                        for length in max_lens
                    ]
                    s = sum(col_widths)
                    col_widths = [w / s for w in col_widths]

                # Largeur de figure basée sur le contenu réel : le
                # plancher était trop haut (6 pouces fixes), ce qui
                # créait un gros espace vide dans les cellules des
                # tableaux à peu de colonnes (ex : "Contacts selon
                # vent", 2 colonnes) — les colonnes remplissaient
                # quand même toute cette largeur inutile puisque
                # leurs proportions sont exprimées en fractions.
                # Un plancher plus bas laisse les tableaux étroits
                # rester compacts, tout en gardant assez de place
                # pour les colonnes à texte long grâce au terme
                # proportionnel au contenu.
                # Hauteur réduite (le titre a été retiré du bloc de
                # calcul ci-dessous, l'ancienne marge haute prévue
                # pour lui n'est donc plus nécessaire) : ça laisse
                # davantage de place à la légende Word ("Tableau N
                # : ...") sur la même page.
                # Hauteur au plus proche du besoin réel : demander
                # une figure plus grande que nécessaire ne fonctionne
                # pas comme on pourrait s'y attendre ici — le tableau
                # matplotlib a une hauteur "naturelle" fixe (liée au
                # nombre de lignes) et reste centré dedans, et
                # bbox_inches="tight" ne recadre pas parfaitement cet
                # excédent (vérifié empiriquement : jusqu'à ~12% de
                # blanc en haut ET en bas avec l'ancienne formule,
                # contre <2% avec celle-ci).
                fig_w = max(2.8, 0.13 * sum(max_lens), 1.1 * ncols)
                fig_h = max(1.2, 0.335 * nrows + 0.15)

                fig, ax = plt.subplots(figsize=(fig_w, fig_h))
                ax.axis("off")

                tbl = ax.table(
                    cellText=rows,
                    colLabels=headers,
                    loc="center",
                    cellLoc="center",
                    colWidths=col_widths
                )
                tbl.auto_set_font_size(False)
                tbl.set_fontsize(9)
                tbl.scale(1, 1.5)

                for j in range(ncols):
                    cell = tbl[(0, j)]
                    cell.set_facecolor("#4C6B3A")
                    cell.set_text_props(color="white", fontweight="bold")

                for i in range(1, nrows + 1):
                    bg = "#F2F5EC" if i % 2 == 0 else "white"
                    for j in range(ncols):
                        tbl[(i, j)].set_facecolor(bg)

                path = os.path.join(
                    out_dir, f"{safe_park}_{year}_T{n:02d}_{name}.jpg"
                )
                fig.savefig(path, dpi=150, bbox_inches="tight")
                plt.close(fig)
                saved.append(path)

            # ---- 1. Espèce / Contacts ----
            tab = (
                df.groupby(species_col).size()
                .reset_index(name="Contacts")
                .sort_values("Contacts", ascending=False)
            )
            save_table(
                ["Espèce", "Contacts"],
                [[str(s), int(c)] for s, c in tab.values],
                "Contacts par espèce",
                "Contacts_par_espece"
            )

            self.pb.setValue(30)

            if date_col:

                # ---- 2. Mois / Contacts ----
                monthly = (
                    df.dropna(subset=[date_col])
                    .groupby(df[date_col].dt.month).size()
                    .reset_index(name="Contacts")
                )
                save_table(
                    ["Mois", "Contacts"],
                    [
                        [mois_fr.get(int(m), str(m)), int(c)]
                        for m, c in monthly.values
                    ],
                    "Contacts par mois",
                    "Contacts_par_mois"
                )

                # ---- 3. Espèce x Mois (pivot + Total) ----
                tmp = df.dropna(subset=[date_col]).copy()
                tmp["Mois"] = tmp[date_col].dt.month
                pivot = tmp.pivot_table(
                    index=species_col, columns="Mois",
                    aggfunc="size", fill_value=0
                )
                months_present = sorted(pivot.columns)
                headers3 = ["Espèce"] + [
                    mois_fr.get(m, str(m)) for m in months_present
                ] + ["Total"]
                rows3 = []
                for sp, row in pivot.iterrows():
                    vals = [int(row[m]) for m in months_present]
                    rows3.append([str(sp)] + vals + [sum(vals)])
                save_table(
                    headers3, rows3,
                    "Contacts par espèce et mois",
                    "Contacts_par_espece_et_mois"
                )

                # ---- 4. Tableau des indicateurs ----
                mois_map7 = {
                    4: "Avril", 5: "Mai", 6: "Juin", 7: "Juil",
                    8: "Août", 9: "Sept", 10: "Oct", 11: "Nov"
                }
                dfx = df.copy()
                dfx["Mois"] = dfx[date_col].dt.month

                lignes = [
                    "Nb contacts chauves-souris",
                    "Nb nuits fonctionnement Barcorder",
                    "Nb nuits positives",
                    "Ratio nuits positives / nuits fonctionnement (%)",
                ]
                data_per_month = {m: [] for m in range(4, 12)}

                for m in range(4, 12):
                    sub = dfx[dfx["Mois"] == m]
                    contacts = len(sub)
                    nuits = calendar.monthrange(year, m)[1]
                    positives = sub[date_col].dt.date.nunique()
                    ratio = round(positives / nuits * 100, 2) if nuits else 0
                    data_per_month[m] = [contacts, nuits, positives, ratio]

                headers4 = ["Indicateur"] + [
                    mois_map7[m] for m in range(4, 12)
                ] + ["Total"]
                rows4 = []
                for i, label in enumerate(lignes):
                    row_vals = [data_per_month[m][i] for m in range(4, 12)]
                    total = round(sum(row_vals), 2) if i == 3 else sum(row_vals)
                    rows4.append([label] + row_vals + [total])
                save_table(
                    headers4, rows4,
                    "Indicateurs par mois",
                    "Indicateurs"
                )

            self.pb.setValue(65)

            # ---- 5. Vent bins ----
            if wind_col:
                v = clean_numeric(df[wind_col]).dropna()
                if not v.empty:
                    bins = [x / 2 for x in range(0, 25)]
                    grp = pd.cut(v, bins=bins, right=False).value_counts().sort_index()
                    total_v = grp.sum()
                    cumul_v = grp.cumsum()
                    save_table(
                        [
                            "Vent (m/s)", "Contacts",
                            "Proportion de contacts cumulés"
                        ],
                        [
                            [
                                f"{it.left:g}-{it.right:g}", int(c),
                                (
                                    f"{100 * cumul_v[it] / total_v:.2f}"
                                    .replace(".", ",") + "%"
                                    if total_v else "0,00%"
                                )
                            ]
                            for it, c in grp.items()
                        ],
                        "Contacts selon vent",
                        "Contacts_selon_vent"
                    )

            # ---- 6. Temp bins ----
            if temp_col:
                t = clean_numeric(df[temp_col]).dropna()
                if not t.empty:
                    bins = list(range(int(t.min()), int(t.max()) + 2))
                    grp = pd.cut(t, bins=bins, right=False).value_counts().sort_index()
                    total_t = grp.sum()
                    cumul_t = grp.cumsum()
                    save_table(
                        [
                            "Température (°C)", "Contacts",
                            "Proportion de contacts cumulés"
                        ],
                        [
                            [
                                f"{int(it.left)}-{int(it.right)}", int(c),
                                (
                                    f"{100 * cumul_t[it] / total_t:.2f}"
                                    .replace(".", ",") + "%"
                                    if total_t else "0,00%"
                                )
                            ]
                            for it, c in grp.items()
                        ],
                        "Contacts selon température",
                        "Contacts_selon_temperature"
                    )

            self.pb.setValue(100)

            # ---- 7. Résumé chiffré pour Albat Rapport ----
            # Petit fichier JSON, à côté des images, qui permet au
            # module Rapport de remplir automatiquement certaines
            # phrases (ex : "X contacts ont été enregistrés...",
            # II.1) sans dupliquer ici la logique de génération de
            # rapport. Ne bloque pas l'export des images en cas
            # d'échec (ex : dossier en lecture seule).
            try:
                if date_col:
                    dates_valides = df[date_col].dropna()
                else:
                    dates_valides = pd.Series([], dtype="datetime64[ns]")

                if not dates_valides.empty:
                    date_premier = dates_valides.min()
                    date_dernier = dates_valides.max()
                    nb_nuits_positives = dates_valides.dt.date.nunique()
                    nb_mois_couverts = (
                        dates_valides.dt.to_period("M").nunique()
                    )
                    nb_jours_couverts = (
                        date_dernier.date() - date_premier.date()
                    ).days + 1
                    date_premier_str = date_premier.date().isoformat()
                    date_dernier_str = date_dernier.date().isoformat()
                else:
                    nb_nuits_positives = 0
                    nb_mois_couverts = 0
                    nb_jours_couverts = None
                    date_premier_str = None
                    date_dernier_str = None

                resume = {
                    "park_name": park_part,
                    "year": year,
                    "total_contacts": int(len(df)),
                    "nb_nuits_positives": int(nb_nuits_positives),
                    "nb_mois_couverts": int(nb_mois_couverts),
                    "date_premier_contact": date_premier_str,
                    "date_dernier_contact": date_dernier_str,
                    "nb_jours_couverts_estime": nb_jours_couverts,
                    # Détail par espèce (nom tel qu'il apparaît dans
                    # le fichier source -> nombre de contacts),
                    # trié du plus au moins contacté. Permet à
                    # Rapport de générer automatiquement la liste
                    # des espèces et d'identifier l'espèce
                    # majoritaire en II.2.
                    "especes": {
                        str(row[species_col]): int(row["Contacts"])
                        for _, row in tab.iterrows()
                    },
                }

                with open(
                    os.path.join(out_dir, "resume_rapport.json"),
                    "w", encoding="utf-8"
                ) as f:
                    json.dump(resume, f, ensure_ascii=False, indent=2)

            except Exception:
                pass

            QMessageBox.information(
                self, "Succès",
                f"{len(saved)} image(s) enregistrée(s) dans :\n{out_dir}"
            )

        except Exception as e:
            QMessageBox.critical(
                self, "Erreur",
                f"Impossible d'exporter les tableaux :\n{e}"
            )

    # ======================================================

    def _confirm_overwrite_file(self, path):
        """
        Si le fichier indiqué existe déjà, demande confirmation
        avant de l'écraser. Retourne True pour continuer (fichier
        absent, ou écrasement confirmé), False pour annuler.
        """

        if not os.path.exists(path):
            return True

        reply = QMessageBox.question(
            self, "Fichier existant",
            f"Le fichier suivant existe déjà :\n{path}\n\n"
            "Voulez-vous l'écraser ?",
            QMessageBox.Yes | QMessageBox.No
        )

        return reply == QMessageBox.Yes

    # ======================================================

    def _confirm_overwrite_batch(self, out_dir, prefix):
        """
        Vérifie si des fichiers correspondant au préfixe donné
        (parc + année) existent déjà dans le dossier de destination
        — signe d'un export précédent — et demande une seule
        confirmation groupée avant de tout écraser. Retourne True
        pour continuer, False pour annuler cette étape.
        """

        try:
            existing = [
                f for f in os.listdir(out_dir)
                if f.startswith(prefix)
            ]
        except Exception:
            existing = []

        if not existing:
            return True

        reply = QMessageBox.question(
            self, "Fichiers existants",
            f"{len(existing)} fichier(s) issus d'un export "
            f"précédent existent déjà dans :\n{out_dir}\n\n"
            "Voulez-vous les écraser ?",
            QMessageBox.Yes | QMessageBox.No
        )

        return reply == QMessageBox.Yes

    # ======================================================

    def generate_all(self):
        """
        Enchaîne les trois exports (rapport Excel, graphiques,
        tableaux) l'un après l'autre, pour éviter de cliquer sur
        chaque bouton séparément. Écrit directement dans le dossier
        'Graph' du projet (default_save_dir), sans redemander de
        dossier à chaque étape — seule une confirmation
        d'écrasement est demandée si des fichiers existent déjà.
        Chaque étape garde sa propre gestion d'erreurs (aucune
        étape ne bloque les suivantes si elle échoue).
        """

        if not self.default_save_dir or not os.path.isdir(
            self.default_save_dir
        ):
            QMessageBox.critical(
                self, "Dossier du projet introuvable",
                "Le dossier 'Graph' du projet n'a pas été trouvé.\n"
                "Ouvrez ou créez un projet depuis l'écran de "
                "démarrage avant d'utiliser 'Tout générer'."
            )
            return

        self.run(target_dir=self.default_save_dir)
        self.export_graphs_images(target_dir=self.default_save_dir)
        self.export_tables_images(target_dir=self.default_save_dir)

    # ======================================================

    def run(self, target_dir=None):

        try:

            if self.city_lat is None or self.city_lon is None:
                raise Exception(
                    "Recherchez et validez une ville de référence "
                    "avant de générer le rapport."
                )

            city_name = self.city_name_found
            lat, lon = self.city_lat, self.city_lon

            if not self.file:
                raise Exception(
                    "Sélectionnez un fichier."
                )

            self.pb.setValue(10)

            df = smart_read_excel(
                self.file
            )

            species_col = find_column(
                df, "species"
            )

            date_col = find_column(
                df, "date"
            )

            time_col = find_column(
                df, "time"
            )

            temp_col = find_column(
                df, "temp"
            )

            wind_col = find_column(
                df, "wind"
            )

            if not species_col:
                raise Exception(
                    "Colonne espèce introuvable."
                )

            year = 2025

            if date_col:

                df[date_col] = pd.to_datetime(
                    df[date_col],
                    errors="coerce"
                )

                if not df[
                    date_col
                ].dropna().empty:

                    year = int(
                        df[date_col]
                        .dt.year
                        .mode()
                        .iloc[0]
                    )

            if self.ref_year:
                year = self.ref_year

            park_part = self.park_name if self.park_name else "Albat"

            if target_dir:
                out = os.path.join(
                    target_dir, f"{park_part}_{year}_Graph.xlsx"
                )
                if not self._confirm_overwrite_file(out):
                    return
            else:
                if self.default_save_dir and os.path.isdir(self.default_save_dir):
                    default_name = os.path.join(
                        self.default_save_dir,
                        f"{park_part}_{year}_Graph.xlsx"
                    )
                else:
                    default_name = str(
                        Path(self.file).with_name(
                            f"{park_part}_{year}_Graph.xlsx"
            )
        )

                out, _ = QFileDialog.getSaveFileName(
                    self,
                    "Enregistrer le rapport",
                    default_name,
                    "Rapport Albat graph (*.xlsx)"
    )

                if not out:
                    return

            if not out.lower().endswith(".xlsx"):
                out += ".xlsx"

            wb = Workbook()

            fill = PatternFill(
                "solid",
                fgColor="D9E1F2"
            )

            # =====================================
            # ONGLET ESPECES
            # =====================================

            ws = wb.active
            ws.title = "espèces"

            ws["A1"] = "Espèce"
            ws["B1"] = "Contacts"

            for c in ["A1", "B1"]:
                ws[c].font = Font(
                    bold=True
                )
                ws[c].fill = fill

            park_part_label = self.park_name if self.park_name else "-"

            ws["D1"] = "Parc éolien"
            ws["E1"] = park_part_label
            ws["D2"] = "Année"
            ws["E2"] = year

            for c in ["D1", "D2"]:
                ws[c].font = Font(bold=True)
                ws[c].fill = fill

            tab = (
                df.groupby(
                    species_col
                )
                .size()
                .reset_index(
                    name="Contacts"
                )
                .sort_values(
                    "Contacts",
                    ascending=False
                )
            )

            r = 2

            for _, row in tab.iterrows():

                ws.cell(
                    r, 1,
                    row.iloc[0]
                )

                ws.cell(
                    r, 2,
                    int(row.iloc[1])
                )

                r += 1

            chart = BarChart()
            chart.style = 2
            chart.title = "Contacts par espèce"
            chart.y_axis.delete = False
            chart.x_axis.delete = False
            chart.y_axis.majorGridlines = None
            chart.y_axis.majorTickMark = "out"
            

            chart.height = 8
            chart.width = 14
            chart.legend = None

            chart.add_data(
                Reference(
                    ws,
                    min_col=2,
                    min_row=1,
                    max_row=r-1
                ),
                titles_from_data=True
            )

            chart.set_categories(
                Reference(
                    ws,
                    min_col=1,
                    min_row=2,
                    max_row=r-1
                )
            )

            ws.add_chart(
                chart,
                "E2"
            )

            # =====================================
            # ONGLET CONTACTS DATE
            # =====================================

            if date_col:

                ws2 = wb.create_sheet(
                    "contacts date"
                )

                ws2["A1"] = "Date"
                ws2["B1"] = "Contacts"

                for c in ["A1", "B1"]:
                    ws2[c].font = Font(
                        bold=True
                    )
                    ws2[c].fill = fill

                daily = (
                    df.dropna(
                        subset=[date_col]
                    )
                    .groupby(
                        df[date_col].dt.date
                    )
                    .size()
                    .reset_index(
                        name="Contacts"
                    )
                )

                rr = 2

                for _, row in daily.iterrows():

                    ws2.cell(
                        rr, 1,
                        str(row.iloc[0])
                    )

                    ws2.cell(
                        rr, 2,
                        int(row.iloc[1])
                    )

                    rr += 1

                line = LineChart()
                line.title = (
                    "Contacts par date"
                )
                line.y_axis.delete = False
                line.x_axis.delete = False
                line.y_axis.majorGridlines = None
                line.y_axis.majorTickMark = "out"

                line.height = 8
                line.width = 14
                line.legend = None

                line.add_data(
                    Reference(
                        ws2,
                        min_col=2,
                        min_row=1,
                        max_row=rr-1
                    ),
                    titles_from_data=True
                )

                line.set_categories(
                    Reference(
                        ws2,
                        min_col=1,
                        min_row=2,
                        max_row=rr-1
                    )
                )

                ws2.add_chart(
                    line,
                    "K2"
                )

                # ---------------------------
                # tableau mensuel
                # ---------------------------

                monthly = (
                    df.dropna(
                        subset=[date_col]
                    )
                    .groupby(
                        df[date_col]
                        .dt.month
                    )
                    .size()
                    .reset_index(
                        name="Contacts"
                    )
                )

                mois_fr = {
                    1:"Janvier",
                    2:"Février",
                    3:"Mars",
                    4:"Avril",
                    5:"Mai",
                    6:"Juin",
                    7:"Juillet",
                    8:"Août",
                    9:"Septembre",
                    10:"Octobre",
                    11:"Novembre",
                    12:"Décembre"
                }

                ws2["H1"] = "Mois"
                ws2["I1"] = "Contacts"

                for c in [
                    "H1", "I1"
                ]:
                    ws2[c].font = Font(
                        bold=True
                    )
                    ws2[c].fill = fill

                mm = 2

                for _, row in monthly.iterrows():

                    ws2.cell(
                        mm, 8,
                        mois_fr.get(
                            int(row.iloc[0]),
                            str(row.iloc[0])
                        )
                    )

                    ws2.cell(
                        mm, 9,
                        int(row.iloc[1])
                    )

                    mm += 1

                bar = BarChart()
                bar.title = (
                    "Contacts par mois"
                )
                bar.y_axis.delete = False
                bar.x_axis.delete = False
                bar.y_axis.majorGridlines = None
                bar.y_axis.majorTickMark = "out"
                bar.height = 7
                bar.width = 12
                bar.legend = None

                bar.add_data(
                    Reference(
                        ws2,
                        min_col=9,
                        min_row=1,
                        max_row=mm-1
                    ),
                    titles_from_data=True
                )

                bar.set_categories(
                    Reference(
                        ws2,
                        min_col=8,
                        min_row=2,
                        max_row=mm-1
                    )
                )

                ws2.add_chart(
                    bar,
                    "E22"
                )

            # =====================================
            # PARTIE 2 CONTINUERA ICI
            # =====================================
# =====================================================
# ALBAT GRAPH V5 METIER COMPLETE
# PARTIE 2 / 3
# À COLLER DANS run(self)
# juste avant :
# wb.save(out)
# =====================================================

            # =====================================
            # ONGLET ESPECES DATE
            # =====================================

            if date_col:

                ws3 = wb.create_sheet(
                    "espèces date"
                )

                tmp = df.dropna(
                    subset=[date_col]
                ).copy()

                tmp["Mois"] = (
                    tmp[date_col]
                    .dt.month
                )

                mois_fr = {
                    1:"Janvier",
                    2:"Février",
                    3:"Mars",
                    4:"Avril",
                    5:"Mai",
                    6:"Juin",
                    7:"Juillet",
                    8:"Août",
                    9:"Septembre",
                    10:"Octobre",
                    11:"Novembre",
                    12:"Décembre"
                }

                pivot = tmp.pivot_table(
                    index=species_col,
                    columns="Mois",
                    aggfunc="size",
                    fill_value=0
                )

                ws3["A1"] = "Espèce"

                for i in range(1, 13):
                    ws3.cell(
                        1, i+1,
                        mois_fr[i]
                    )

                ws3.cell(
                    1, 14,
                    "Total"
                )

                rr = 2

                for idx, row in pivot.iterrows():

                    ws3.cell(
                        rr, 1,
                        str(idx)
                    )

                    total = 0

                    for i in range(1, 13):

                        val = (
                            int(row[i])
                            if i in pivot.columns
                            else 0
                        )

                        ws3.cell(
                            rr,
                            i+1,
                            val
                        )

                        total += val

                    ws3.cell(
                        rr, 14,
                        total
                    )

                    rr += 1

                chart3 = BarChart()
                chart3.title = (
                    "Contacts par espèce et mois"
                )

                chart3.height = 8
                chart3.width = 16
                chart3.y_axis.delete = False
                chart3.x_axis.delete = False
                chart3.y_axis.majorGridlines = None
                chart3.y_axis.majorTickMark = "out"
                chart3.legend.overlay = False
                chart3.legend.position = "b"

                chart3.add_data(
                    Reference(
                        ws3,
                        min_col=1,
                        max_col=13,
                        min_row=2,
                        max_row=rr-1
                    ),
                    from_rows=True,
                    titles_from_data=True
                )

                chart3.set_categories(
                    Reference(
                        ws3,
                        min_col=2,
                        max_col=13,
                        min_row=1,
                        max_row=1
                    )
                )

                ws3.add_chart(
                    chart3,
                    "P2"
                )
            # =====================================
            # TABLEAU REGROUPE PAR TAXONS
            # =====================================

            ws3["A19"] = "Groupe"
            ws3["B19"] = "Avril"
            ws3["C19"] = "Mai"
            ws3["D19"] = "Juin"
            ws3["E19"] = "Juillet"
            ws3["F19"] = "Août"
            ws3["G19"] = "Septembre"
            ws3["H19"] = "Octobre"
            ws3["I19"] = "Novembre"
            

            for c in range(1, 15):
                ws3.cell(1, c).font = Font(bold=True)
                ws3.cell(1, c).fill = fill

            for c in range(1, 10):
                ws3.cell(19, c).font = Font(bold=True)
                ws3.cell(19, c).fill = fill

            # noms groupes
            ws3["A20"] = "Noctule + Sérotine"
            ws3["A21"] = "Pipistrelles"

            # dictionnaire colonnes mois
            col_map = {
                4: 2,
                5: 3,
                6: 4,
                7: 5,
                8: 6,
                9: 7,
                10: 8,
                11: 9
            }

            for mois in range(4, 12):

                sub = tmp[tmp["Mois"] == mois]

                # groupe 1
                g1 = sub[
                    sub[species_col].str.contains(
                        "Noctule|Sérotine|Serotine",
                        case=False,
                        na=False
                    )
                ]

                # groupe 2
                g2 = sub[
                    sub[species_col].str.contains(
                        "Pipistrelle|Pipistrelles",
                        case=False,
                        na=False
                    )
                ]

                col = col_map[mois]

                ws3.cell(20, col, len(g1))
                ws3.cell(21, col, len(g2))

           

            # =====================================
            # HISTOGRAMME GROUPES TAXONOMIQUES
            # =====================================


            chart_grp = BarChart()

            chart_grp.type = "col"
            chart_grp.style = 10
            chart_grp.title = "Activité Noctule / Sérotine / Pipistrelles"

            chart_grp.height = 8
            chart_grp.width = 14

            chart_grp.y_axis.delete = False
            chart_grp.x_axis.delete = False
            chart_grp.y_axis.majorGridlines = None
            chart_grp.y_axis.majorTickMark = "out"

            # série 1
            data1 = Reference(
                ws3,
                min_col=2,
                max_col=9,
                min_row=20,
                max_row=20
            )
          
            # série 2
            data2 = Reference(
                ws3,
                min_col=2,
                max_col=9,
                min_row=21,
                max_row=21
            )

            serie1 = Series(data1, title="Noctule + Sérotine")
            serie2 = Series(data2, title="Pipistrelles")

            chart_grp.series.append(serie1)
            chart_grp.series.append(serie2)

            # mois
            cats = Reference(
                ws3,
                min_col=2,
                max_col=9,
                min_row=19,
                max_row=19
            )

            chart_grp.set_categories(cats)

            ws3.add_chart(chart_grp, "P20")  
            # =====================================
            # ONGLET VENT
            # =====================================

            if wind_col:

                ws4 = wb.create_sheet(
                    "vent"
                )

                ws4["A1"] = "Vent"
                ws4["B1"] = "Contacts"

                for c in ["A1", "B1"]:
                    ws4[c].font = Font(
                        bold=True
                    )
                    ws4[c].fill = fill

                v = clean_numeric(
                    df[wind_col]
                ).dropna()

                bins = [
                    x/2 for x in range(0, 25)
                ]

                grp = (
                    pd.cut(
                        v,
                        bins=bins,
                        right=False
                    )
                    .value_counts()
                    .sort_index()
                )

                rr = 2

                for inter, val in grp.items():

                    ws4.cell(
                        rr, 1,
                        f"{inter.left:g}-{inter.right:g}"
                    )

                    ws4.cell(
                        rr, 2,
                        int(val)
                    )

                    rr += 1

                chart4 = BarChart()
                chart4.style = 5
                chart4.title = (
                    "Contacts selon vent"
                )
                chart4.y_axis.delete = False
                chart4.x_axis.delete = False
                chart4.y_axis.majorGridlines = None
                chart4.y_axis.majorTickMark = "out"

                chart4.height = 8
                chart4.width = 14
                chart4.legend = None

                chart4.add_data(
                    Reference(
                        ws4,
                        min_col=2,
                        min_row=1,
                        max_row=rr-1
                    ),
                    titles_from_data=True
                )

                chart4.set_categories(
                    Reference(
                        ws4,
                        min_col=1,
                        min_row=2,
                        max_row=rr-1
                    )
                )

                ws4.add_chart(
                    chart4,
                    "E2"
                )

            # =====================================
            # ONGLET TEMP
            # =====================================

            if temp_col:

                ws5 = wb.create_sheet(
                    "temp"
                )

                ws5["A1"] = "Température"
                ws5["B1"] = "Contacts"

                for c in ["A1", "B1"]:
                    ws5[c].font = Font(
                        bold=True
                    )
                    ws5[c].fill = fill

                t = clean_numeric(
                    df[temp_col]
                ).dropna()

                bins = list(
                    range(
                        int(t.min()),
                        int(t.max()) + 2
                    )
                )

                grp = (
                    pd.cut(
                        t,
                        bins=bins,
                        right=False
                    )
                    .value_counts()
                    .sort_index()
                )

                rr = 2

                for inter, val in grp.items():

                    ws5.cell(
                        rr, 1,
                        f"{int(inter.left)}-{int(inter.right)}"
                    )

                    ws5.cell(
                        rr, 2,
                        int(val)
                    )

                    rr += 1

                chart5 = BarChart()
                chart5.style = 3
                chart5.title = (
                    "Contacts selon température"
                )
                chart5.y_axis.delete = False
                chart5.x_axis.delete = False
                chart5.y_axis.majorGridlines = None
                chart5.y_axis.majorTickMark = "out"

                chart5.height = 8
                chart5.width = 14
                chart5.legend = None

                chart5.add_data(
                    Reference(
                        ws5,
                        min_col=2,
                        min_row=1,
                        max_row=rr-1
                    ),
                    titles_from_data=True
                )

                chart5.set_categories(
                    Reference(
                        ws5,
                        min_col=1,
                        min_row=2,
                        max_row=rr-1
                    )
                )

                ws5.add_chart(
                    chart5,
                    "E2"
                )

            # =====================================
            # PARTIE 3 CONTINUERA ICI
            # =====================================
# =====================================================
# ALBAT GRAPH V5 METIER COMPLETE
# PARTIE 3 / 3
# À COLLER DANS run(self)
# juste avant wb.save(out)
# =====================================================
# =====================================
# ONGLET NB HEURES DEFINITIF
# Corrige H+23 / passage minuit
# =====================================

            if date_col and time_col:

                ws6 = wb.create_sheet("nb heures")
                ws6["D1"] = "Ville de référence"
                ws6["E1"] = city_name

                ws6["D1"].font = Font(bold=True)
                ws6["D1"].fill = fill

                ws6["D2"] = "Parc éolien"
                ws6["E2"] = self.park_name if self.park_name else "-"
                ws6["D3"] = "Année"
                ws6["E3"] = year

                for c in ["D2", "D3"]:
                    ws6[c].font = Font(bold=True)
                    ws6[c].fill = fill

                ws6["A1"] = "Heures après coucher soleil"
                ws6["B1"] = "Contacts"

                for c in ["A1", "B1"]:
                    ws6[c].font = Font(bold=True)
                    ws6[c].fill = fill

                tmp = df.copy()

                tmp[date_col] = pd.to_datetime(
                    tmp[date_col],
                    errors="coerce"
                )

                tmp[time_col] = pd.to_datetime(
                    tmp[time_col].astype(str),
                    errors="coerce"
                )

                lignes = []

                for _, row in tmp.iterrows():

                    if pd.isna(row[date_col]) or pd.isna(row[time_col]):
                        continue

                    d = row[date_col]

                    heure_contact = datetime(
                        d.year,
                        d.month,
                        d.day,
                        row[time_col].hour,
                        row[time_col].minute,
                        row[time_col].second
                    )

                    sunset = calc_sunset(
                        d,
                        lat,
                        lon
                    )

                    if sunset is None:
                        continue

                    # ==========================
                    # PASSAGE MINUIT
                    # Si heure entre 00h et 11h59
                    # on considère nuit suivante
                    # ==========================
                    if heure_contact.hour < 12:
                        heure_contact += timedelta(days=1)

                    delta = heure_contact - sunset

                    heures = delta.total_seconds() / 3600

                    # garde seulement vraie activité nocturne
                    if heures < 0:
                        continue

                    # limite réaliste
                    if heures > 12:
                        continue

                    classe = int(heures)

                    lignes.append(classe)

                # ==========================
                # Tableau complet H+0 à H+12
                # ==========================
                compte = {i: 0 for i in range(13)}

                for x in lignes:
                    if x in compte:
                        compte[x] += 1

                rr = 2

                for hr in range(13):

                    ws6.cell(rr, 1, f"H+{hr}")
                    ws6.cell(rr, 2, compte[hr])

                    rr += 1

                chart6 = BarChart()
                chart6.title = "Contacts après coucher du soleil"
                chart6.height = 8
                chart6.width = 14
                chart6.legend = None
                chart6.y_axis.delete = False
                chart6.x_axis.delete = False
                chart6.y_axis.majorGridlines = None
                chart6.y_axis.majorTickMark = "out"

                chart6.add_data(
                    Reference(
                        ws6,
                        min_col=2,
                        min_row=1,
                        max_row=rr-1
                    ),
                    titles_from_data=True
                )

                chart6.set_categories(
                    Reference(
                        ws6,
                        min_col=1,
                        min_row=2,
                        max_row=rr-1
                    )
                )

                ws6.add_chart(chart6, "G2")
            # =====================================
            # ONGLET TABLEAU
            # =====================================

            if date_col:

                    ws7 = wb.create_sheet(
                        "tableau"
                    )

                    mois_map = {
                        4:"Avril",
                        5:"Mai",
                        6:"Juin",
                        7:"Juil",
                        8:"Août",
                        9:"Sept",
                        10:"Oct",
                        11:"Nov"
                    }

                    ws7["A1"] = "Indicateur"

                    col = 2

                    for m in range(4, 12):

                        ws7.cell(
                            1, col,
                            mois_map[m]
                        )

                        col += 1

                    ws7.cell(
                        1, col,
                        "Total"
                    )

                    lignes = [
                        "Nb contacts chauves-souris",
                        "Nb nuits fonctionnement Barcorder",
                        "Nb nuits positives",
                        "Ratio nuits positives / nuits fonctionnement"
                    ]

                    for i, txt in enumerate(
                        lignes,
                        start=2
                    ):
                        ws7.cell(
                            i, 1, txt
                        )

                    for c in range(1, 11):

                        ws7.cell(
                            1, c
                        ).font = Font(
                            bold=True
                        )

                        ws7.cell(
                            1, c
                        ).fill = fill

                    dfx = df.copy()

                    dfx["Mois"] = (
                        dfx[date_col]
                        .dt.month
                    )

                    cc = 2

                    for m in range(4, 12):

                        sub = dfx[
                            dfx["Mois"] == m
                        ]

                        contacts = len(sub)

                        nuits = calendar.monthrange(
                            year, m
                        )[1]

                        positives = (
                            sub[date_col]
                            .dt.date
                            .nunique()
                        )

                        ratio = round(
                            (
                                positives / nuits * 100
                            ),
                            2
                        ) if nuits else 0

                        ws7.cell(
                            2, cc,
                            contacts
                        )

                        ws7.cell(
                            3, cc,
                            nuits
                        )

                        ws7.cell(
                            4, cc,
                            positives
                        )

                        ws7.cell(
                            5, cc,
                            ratio
                        )

                        cc += 1

                    for r in range(2, 5):

                        ws7.cell(
                            r, 10,
                            f'=SUM(B{r}:I{r})'
                        )

            # =====================================
            # FINAL SAVE
            # =====================================

            self.pb.setValue(100)
            for ws in wb.worksheets:
                autofit_worksheet(ws)

            wb.save(out)

            open_file(out)

            play_success_sound()

            QMessageBox.information(
                self,
                "Succès",
                f"Rapport créé : {out}"
            
            
)
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur",
                str(e)
            )

if __name__ == "__main__":
    app = QApplication(sys.argv)

    win = GraphWindow()
    win.show()

    sys.exit(app.exec())