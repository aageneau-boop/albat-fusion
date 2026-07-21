# ==========================================================
# ALBAT BRIDAGE V2
# Version corrigée
# ==========================================================

import sys
import os
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from PySide6.QtWidgets import (
    QApplication,
    QWidget,
    QLabel,
    QPushButton,
    QFileDialog,
    QInputDialog,
    QVBoxLayout,
    QHBoxLayout,
    QMessageBox,
    QProgressBar,
    QFrame,
    QSpinBox,
    QDateEdit,
    QScrollArea,
    QLineEdit,
    QCheckBox,
    QComboBox,
    QGraphicsDropShadowEffect

)

from PySide6.QtCore import Qt, QDate, QPoint, QTimer
from PySide6.QtGui import ( QPixmap, QIcon, QFontDatabase, QFont )
from modules.utils import resource_path

import datetime as dtmod
from zoneinfo import ZoneInfo

try:
    from astral import LocationInfo
    from astral.geocoder import database as astral_database, lookup as astral_lookup
    from astral.sun import sun as astral_sun
    ASTRAL_OK = True
except ImportError:
    ASTRAL_OK = False


# ==================================================
# ASSETS
# ==================================================
ASSETS = resource_path("assets")

BG = os.path.join(ASSETS, "background_graph.png")
APP_ICON = os.path.join(ASSETS, "icone_albat.icns")

if not os.path.exists(APP_ICON):
    APP_ICON = os.path.join(ASSETS, "icone_albat.ico")



# ==========================================
# COMPATIBILITE MAC / RETINA
# ==========================================
if sys.platform == "darwin":
    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
    os.environ["QT_ENABLE_HIGHDPI_SCALING"] = "1"


class BridageWindow(QWidget):

    def __init__(self, park_name="", ref_city="", ref_year=None, bureau_etude="", default_save_dir="", correlations_dir_hint=""):
        super().__init__()

        self.park_name = park_name
        self.ref_city = ref_city
        self.ref_year = ref_year
        self.bureau_etude = bureau_etude
        self.default_save_dir = default_save_dir
        self.correlations_dir = correlations_dir_hint

        font_path = os.path.join(
            ASSETS,
            "OpenSans-VariableFont_wdth,wght.ttf"
        )

        if os.path.exists(font_path):
            QFontDatabase.addApplicationFont(font_path)

        self.files = []
        self.plans = []
        self.location_info = None

        self.setWindowTitle("Albat Bridage")
        self.resize(500, 900)
        self.setMinimumSize(500, 900)

        if os.path.exists(APP_ICON):
            self.setWindowIcon(QIcon(APP_ICON))

        self.setStyleSheet("""
QWidget{
    background:#071107;
    color:white;
    font-family:"Helvetica Neue","Segoe UI",Arial,sans-serif;
}

QFrame{
    background: transparent;
    border-radius:18px;
}

QPushButton{
    background:rgba(90,127,71,145);
    border:1px solid rgba(255,255,255,40);
    border-radius:16px;
    padding:14px;
    font-size:15px;
    font-weight:700;
    color:white;
}

QPushButton:hover{
    background:rgba(130,184,99,220);
}

QLineEdit,
QSpinBox,
QDateEdit{
    background:#081008;
    border-radius:12px;
    padding:8px;
    border:1px solid rgba(255,255,255,35);
    color:#eef7d0;
    font-size:12px;
}

QProgressBar{
    background:#081008;
    border-radius:11px;
    height:22px;
}

QProgressBar::chunk{
    background:#a6f56f;
    border-radius:11px;
}
""")

        self.ui()

        # Pré-remplissage ET validation automatique de la ville de
        # référence transmise depuis le popup de démarrage du
        # projet : évite d'avoir à recliquer sur "Rechercher" à
        # chaque ouverture d'onglet. search_city() gère déjà ses
        # propres erreurs (ville introuvable, pas de connexion) sans
        # lever d'exception : l'onglet reste utilisable si ça échoue,
        # il faudra juste valider la ville manuellement.
        if self.ref_city:
            self.city_input.setText(self.ref_city)
            self.search_city()

        # Auto-chargement des fichiers Excel du dossier
        # "Correlations" du projet, pour ne pas avoir à les
        # sélectionner à nouveau manuellement dans chaque onglet.
        # Bridage accepte plusieurs fichiers à la fois (comme
        # pick_file) : tous les fichiers Excel valides trouvés sont
        # chargés.
        auto_files = self._find_correlations_excels()
        if auto_files:
            self.files = auto_files
            if len(auto_files) == 1:
                self.lbl_file.setText(os.path.basename(auto_files[0]))
            else:
                self.lbl_file.setText(
                    f"{len(auto_files)} fichiers sélectionnés"
                )

    # ======================================================

    def _find_correlations_excels(self):
        """
        Cherche, dans le dossier 'Correlations' du projet, tous les
        fichiers Excel (.xlsx/.xlsm/.xls) à charger automatiquement.
        Ignore les fichiers verrous (~$...) et cachés. Retourne une
        liste (vide si dossier absent ou sans fichier Excel).
        """

        if not self.correlations_dir or not os.path.isdir(
            self.correlations_dir
        ):
            return []

        try:
            fichiers = sorted(os.listdir(self.correlations_dir))
        except Exception:
            return []

        return [
            os.path.join(self.correlations_dir, f)
            for f in fichiers
            if not f.startswith("~$")
            and not f.startswith(".")
            and f.lower().endswith((".xlsx", ".xlsm", ".xls"))
        ]


    def show_calendar(self, date_edit):

        cal = date_edit.calendarWidget()

        cal.setWindowFlag(Qt.Popup, True)

        pos = self.mapToGlobal(QPoint(10, 10))

        cal.move(pos)

        try:
            cal.raise_()
            cal.activateWindow()
        except:
            pass

        cal.show()


    def resizeEvent(self, event):

        self.bg.setGeometry(self.rect())

        if os.path.exists(BG):

            p = QPixmap(BG)

            if not p.isNull():

                self.bg.setPixmap(
                    p.scaled(
                        self.size(),
                        Qt.KeepAspectRatioByExpanding,
                        Qt.SmoothTransformation
                    )
                )

        super().resizeEvent(event)


    # ======================================================
    # UI
    # ======================================================

    def ui(self):

        self.bg = QLabel(self)
        self.bg.setGeometry(self.rect())
        self.bg.setAlignment(Qt.AlignCenter)

        if os.path.exists(BG):

            p = QPixmap(BG)

            if not p.isNull():

                self.bg.setPixmap(
                    p.scaled(
                        self.size(),
                        Qt.KeepAspectRatioByExpanding,
                        Qt.SmoothTransformation
                    )
                )

            self.bg.lower()

        main = QVBoxLayout(self)

        main.setSpacing(0)
        main.setContentsMargins(20, 6, 20, 20)

        # Le grand titre décoratif "Albat / BRIDAGE" a été retiré :
        # l'onglet actif dans la barre d'onglets du haut suffit
        # déjà à indiquer où l'on se trouve.
        main.addSpacing(10)

        self.lbl_file = QLabel(
            "Aucun fichier sélectionné"
        )

        self.lbl_file.setStyleSheet("""
background:rgba(255,255,255,20);
border-radius:12px;
padding:8px;
""")

        btn_file = QPushButton(
            "Choisir fichier Excel"
        )

        btn_file.clicked.connect(
            self.pick_file
        )

        main.addWidget(btn_file, alignment=Qt.AlignHCenter)

        self.lbl_file.setAlignment(Qt.AlignCenter)
        self.lbl_file.setStyleSheet("""
            color: black;
            font-size: 13px;
            font-weight: bold;
        """)
        self.lbl_file.setFixedWidth(420)

        main.addWidget(
            self.lbl_file,
            alignment=Qt.AlignHCenter
        )

        # ======================================================
        # BLOC REFERENCE SOLAIRE (lever / coucher du soleil)
        # ======================================================

        sun_title = QLabel("RÉFÉRENCE SOLAIRE")
        sun_title.setAlignment(Qt.AlignCenter)
        sun_title.setStyleSheet("""
QLabel {
    color: #E6E8C8;
    font-weight: 700;
    font-size: 14px;
    letter-spacing: 1px;
    background-color: rgba(0, 0, 0, 0.50);
    border: 1px solid rgba(140,170,90,0.60);
    border-radius: 12px;
    padding-top: 1px;
    padding-bottom: 1px;
    padding-left: 26px;
    padding-right: 26px;
    margin-top: 4px;
    margin-bottom: 0px;
    min-height: 20px;
    max-height: 24px;
}
""")
        main.addWidget(sun_title, alignment=Qt.AlignHCenter)
        main.addSpacing(6)

        sun_box = QFrame()
        sun_box.setStyleSheet("""
            QFrame{
                background:rgba(45,75,35,215);
                border:1px solid rgba(255,255,255,45);
                border-radius:16px;
            }
        """)

        sun_layout = QVBoxLayout(sun_box)
        sun_layout.setSpacing(8)
        sun_layout.setContentsMargins(14, 12, 14, 12)

        lab_sun = QLabel("Ville de référence (lever / coucher du soleil)")
        lab_sun.setStyleSheet("""
            color: white;
            font-size: 14px;
            font-weight: bold;
            background:transparent;
        """)
        sun_layout.addWidget(lab_sun)

        city_line = QHBoxLayout()

        self.city_input = QLineEdit()
        self.city_input.setPlaceholderText("Ville de référence (ex : Dijon)")

        btn_city = QPushButton("Rechercher")
        btn_city.setStyleSheet("""
            QPushButton{
                background:rgba(120,150,90,180);
                border:1px solid rgba(255,255,255,35);
                border-radius:12px;
                color:white;
                font-size:13px;
                font-weight:700;
                padding:8px 16px;
            }
            QPushButton:hover{
                background:rgba(160,210,120,220);
            }
            QPushButton:pressed{
                background:rgba(90,120,70,255);
            }
        """)
        btn_city.clicked.connect(self.search_city)
        self.city_input.returnPressed.connect(self.search_city)

        city_line.addWidget(self.city_input)
        city_line.addWidget(btn_city)

        sun_layout.addLayout(city_line)

        self.city_status = QLabel(
            "Aucune ville recherchée." if ASTRAL_OK
            else "Module 'astral' manquant : exécutez 'pip install astral'."
        )
        self.city_status.setWordWrap(True)
        self.city_status.setStyleSheet("""
            color: #eef7d0;
            font-size: 11px;
            background:transparent;
        """)
        sun_layout.addWidget(self.city_status)

        note_sun = QLabel(
            "Cette ville sert de référence pour tous les plans qui "
            "activent la restriction nocturne. Chaque plan de bridage "
            "peut ensuite activer/désactiver cette restriction et régler "
            "ses propres décalages (avant/après coucher et lever) "
            "indépendamment des autres plans."
        )
        note_sun.setWordWrap(True)
        note_sun.setStyleSheet("""
            color: #f3f8e2;
            font-size: 10px;
            background: rgba(0,0,0,0.30);
            border-radius: 8px;
            padding: 6px 8px;
        """)
        sun_layout.addWidget(note_sun)

        main.addWidget(sun_box)

        # ── Enregistrer / charger une configuration de plans ──
        save_load_line = QHBoxLayout()
        save_load_line.setSpacing(10)

        save_load_style = """
            QPushButton{
                background:rgba(70,100,55,190);
                border:1px solid rgba(255,255,255,35);
                border-radius:10px;
                color:white;
                font-size:11px;
                font-weight:700;
                padding:10px 6px;
            }
            QPushButton:hover{
                background:rgba(100,140,75,220);
            }
        """

        save_plans_btn = QPushButton("Enregistrer les plans")
        save_plans_btn.setMinimumHeight(34)
        save_plans_btn.setStyleSheet(save_load_style)
        save_plans_btn.clicked.connect(self.save_plans_to_file)

        load_plans_btn = QPushButton("Charger des plans")
        load_plans_btn.setMinimumHeight(34)
        load_plans_btn.setStyleSheet(save_load_style)
        load_plans_btn.clicked.connect(self.load_plans_from_file)

        save_load_line.addWidget(save_plans_btn, 1)
        save_load_line.addWidget(load_plans_btn, 1)

        main.addLayout(save_load_line)

        # Décalage vertical du bloc plans de bridage
        main.addSpacing(8)

        self.plans_container = QVBoxLayout()
        self.plans_container.setSpacing(14)

        plans_title = QLabel("PLANS DE BRIDAGE")
        plans_title.setAlignment(Qt.AlignCenter)
        plans_title.setStyleSheet("""
QLabel {
    color: #E6E8C8;
    font-weight: 700;
    font-size: 14px;
    letter-spacing: 1px;
    background-color: rgba(0, 0, 0, 0.50);
    border: 1px solid rgba(140,170,90,0.60);
    border-radius: 12px;
    padding-top: 1px;
    padding-bottom: 1px;
    padding-left: 26px;
    padding-right: 26px;
    margin-top: 4px;
    margin-bottom: 0px;
    min-height: 20px;
    max-height: 24px;
}
""")

        self.plans_container.addWidget(plans_title, alignment=Qt.AlignHCenter)

        plans_note = QLabel(
            "Chaque plan génère sa propre colonne dans le fichier de "
            "sortie (nommée d'après le plan) et son propre résumé "
            "dans l'onglet Résumé."
        )
        plans_note.setWordWrap(True)
        plans_note.setStyleSheet("""
            color: #cfe0a0;
            font-size: 10px;
            background: transparent;
            padding-left: 4px;
            padding-right: 4px;
        """)
        self.plans_container.addWidget(plans_note)

        scroll_widget = QWidget()
        scroll_widget.setLayout(
            self.plans_container
        )
        scroll_widget.setStyleSheet("background: transparent;")

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setWidget(scroll_widget)
        scroll.setFixedHeight(520)
        scroll.setStyleSheet("""
            QScrollArea {
                background: transparent;
                border: none;
            }
            QScrollArea > QWidget > QWidget {
                background: transparent;
            }
        """)
        scroll.viewport().setStyleSheet("background: transparent;")

        main.addWidget(scroll)

        # Espace réduit pour remonter le bouton
        main.addSpacing(10)

        add_plan_btn = QPushButton(
            "+ Ajouter un plan de bridage"
        )

        add_plan_btn.clicked.connect(
            self.add_plan
        )

        main.addWidget(add_plan_btn)

        self.add_plan()

        self.pb = QProgressBar()

        main.addWidget(self.pb)

        run = QPushButton(
            "GENERER BRIDAGE"
        )

        run.setMinimumHeight(52)

        run.clicked.connect(
            self.run
        )

        main.addWidget(run)

    # ======================================================

    def add_plan(self):

        plan_card = QFrame()
        plan_card.setStyleSheet("""
            QFrame{
                background:rgba(59,41,25,225);
                border:1px solid rgba(255,255,255,40);
                border-radius:16px;
            }
        """)

        plan_layout = QVBoxLayout(plan_card)
        plan_layout.setContentsMargins(12, 10, 12, 12)
        plan_layout.setSpacing(8)

        header_line = QHBoxLayout()
        header_line.setSpacing(8)

        lab_name = QLabel("Nom du plan")
        lab_name.setStyleSheet("""
            color: white;
            font-size: 13px;
            font-weight: bold;
            background: transparent;
        """)

        name_edit = QLineEdit()
        name_edit.setText(f"Plan {len(self.plans) + 1}")

        delete_plan_btn = QPushButton("✕")
        delete_plan_btn.setFixedSize(24, 24)
        delete_plan_btn.setStyleSheet("""
            QPushButton{
                background:rgba(110,140,80,180);
                border:2px solid rgba(255,255,255,40);
                border-radius:12px;
                color:white;
                font-size:12px;
                font-weight:900;
                padding:0px;
            }
            QPushButton:hover{
                background:rgba(140,180,100,220);
            }
            QPushButton:pressed{
                background:rgba(90,120,70,255);
            }
        """)

        header_line.addWidget(lab_name)
        header_line.addWidget(name_edit)
        header_line.addWidget(delete_plan_btn)

        plan_layout.addLayout(header_line)

        period_container = QVBoxLayout()
        period_container.setSpacing(8)
        plan_layout.addLayout(period_container)

        add_period_btn = QPushButton("Ajouter une période à ce plan")
        add_period_btn.setStyleSheet("""
            QPushButton{
                background:rgba(70,100,55,190);
                border:1px solid rgba(255,255,255,35);
                border-radius:10px;
                color:white;
                font-size:12px;
                font-weight:700;
                padding:8px;
            }
            QPushButton:hover{
                background:rgba(100,140,75,220);
            }
        """)
        plan_layout.addWidget(add_period_btn)

        self.plans_container.addWidget(plan_card)

        plan = {
            "card": plan_card,
            "name_edit": name_edit,
            "period_container": period_container,
            "periods": [],
        }

        self.plans.append(plan)

        add_period_btn.clicked.connect(
            lambda _, p=plan: self.add_period(p)
        )

        delete_plan_btn.clicked.connect(
            lambda _, p=plan: self.remove_plan(p)
        )

        # Chaque plan démarre avec une première période pré-remplie
        self.add_period(plan)

    # ======================================================

    def remove_plan(self, plan):

        if len(self.plans) <= 1:
            QMessageBox.warning(
                self,
                "Impossible",
                "Il doit rester au moins un plan de bridage."
            )
            return

        self._remove_plan_forced(plan)

    # ======================================================

    def _remove_plan_forced(self, plan):
        """Supprime un plan sans vérifier qu'il en reste au moins un
        (utilisé en interne, notamment avant de charger une
        configuration de plans depuis un fichier)."""

        card = plan["card"]

        self.plans_container.removeWidget(card)

        card.setParent(None)

        card.deleteLater()

        if plan in self.plans:
            self.plans.remove(plan)

    # ======================================================

    def _collect_plans_data(self):
        """Construit le dictionnaire de données des plans de
        bridage (réutilisé par la sauvegarde manuelle et par la
        sauvegarde globale du projet)."""

        data = {
            "ville_reference": self.city_input.text().strip(),
            "plans": []
        }

        for plan in self.plans:

            plan_data = {
                "nom": plan["name_edit"].text().strip(),
                "periodes": []
            }

            for i, p in enumerate(plan["periods"], start=1):

                plan_data["periodes"].append({
                    "nom": f"Période {i}",
                    "debut": p["start"].date().toString("dd/MM/yyyy"),
                    "fin": p["end"].date().toString("dd/MM/yyyy"),
                    "vent_actif": not p["no_vent_limit"].isChecked(),
                    "vent_max": p["vent"].value(),
                    "temp_actif": not p["no_temp_limit"].isChecked(),
                    "temp_min": p["temp"].value(),
                    "nocturne_actif": p["use_sun"].isChecked(),
                    "reference_debut": (
                        "coucher" if p["start_ref"].currentIndex() == 0
                        else "lever"
                    ),
                    "decalage_debut": p["offset_start"].value(),
                    "reference_fin": (
                        "lever" if p["end_ref"].currentIndex() == 0
                        else "coucher"
                    ),
                    "decalage_fin": p["offset_end"].value(),
                })

            data["plans"].append(plan_data)

        return data

    # ======================================================

    def save_plans_to_path(self, out_path):
        """Enregistre les plans directement vers un chemin donné,
        sans passer par une boîte de dialogue (utilisé par la
        sauvegarde globale du projet)."""

        data = self._collect_plans_data()

        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    # ======================================================

    def save_plans_to_file(self):
        """
        Enregistre la configuration actuelle de tous les plans de
        bridage (noms, périodes, seuils vent/température, restriction
        nocturne et décalages) dans un fichier .json réutilisable.

        Utilise un schéma commun avec Albat Scenar (clé "periodes",
        restriction nocturne par période) pour permettre l'échange
        de fichiers entre les deux modules.
        """

        try:

            default_plans_name = "plans_bridage.json"

            if self.default_save_dir and os.path.isdir(self.default_save_dir):
                default_plans_name = os.path.join(
                    self.default_save_dir, default_plans_name
                )

            out_path, _ = QFileDialog.getSaveFileName(
                self,
                "Enregistrer les plans de bridage",
                default_plans_name,
                "Fichier de plans (*.json)"
            )

            if not out_path:
                return

            if not out_path.lower().endswith(".json"):
                out_path += ".json"

            self.save_plans_to_path(out_path)

            QMessageBox.information(
                self,
                "Succès",
                f"{len(self.plans)} plan(s) enregistré(s) dans :\n{out_path}"
            )

        except Exception as e:
            QMessageBox.critical(
                self, "Erreur",
                f"Impossible d'enregistrer les plans :\n{e}"
            )

    # ======================================================

    def _apply_plans_data(self, data):
        """Applique un dictionnaire de données de plans à
        l'interface (réutilisé par le chargement manuel et par le
        chargement global du projet)."""

        # Détection du format : liste de plans (Bridage) ou
        # un seul scénario/plan avec ses périodes (Scenar).
        if "plans" in data:
            plans_data = data.get("plans", [])
        elif "periodes" in data:
            plans_data = [{
                "nom": data.get("nom") or data.get("nom_scenario", "Plan 1"),
                "periodes": data.get("periodes", []),
            }]
        else:
            plans_data = []

        # Supprime tous les plans actuellement affichés
        for plan in list(self.plans):
            self._remove_plan_forced(plan)

        ville = data.get("ville_reference", "")
        if ville:
            self.city_input.setText(ville)

        for plan_data in plans_data:

            self.add_plan()
            plan = self.plans[-1]

            plan["name_edit"].setText(
                plan_data.get("nom", plan["name_edit"].text())
            )

            # Le plan démarre avec une période par défaut :
            # on la retire, puis on recrée exactement les
            # périodes enregistrées.
            for p in list(plan["periods"]):
                self.remove_period(plan, p)

            for p_data in plan_data.get("periodes", []):

                self.add_period(plan)
                p = plan["periods"][-1]

                debut = QDate.fromString(
                    p_data.get("debut", "01/04/2025"), "dd/MM/yyyy"
                )
                fin = QDate.fromString(
                    p_data.get("fin", "31/10/2025"), "dd/MM/yyyy"
                )

                if debut.isValid():
                    p["start"].setDate(debut)
                if fin.isValid():
                    p["end"].setDate(fin)

                # Compatible schéma actuel (vent_actif/temp_actif)
                # et ancien schéma Bridage (vent_sans_limite/...)
                if "vent_actif" in p_data:
                    p["no_vent_limit"].setChecked(
                        not p_data.get("vent_actif", True)
                    )
                else:
                    p["no_vent_limit"].setChecked(
                        p_data.get("vent_sans_limite", False)
                    )
                p["vent"].setValue(round(p_data.get("vent_max", 6)))

                if "temp_actif" in p_data:
                    p["no_temp_limit"].setChecked(
                        not p_data.get("temp_actif", True)
                    )
                else:
                    p["no_temp_limit"].setChecked(
                        p_data.get("temp_sans_limite", False)
                    )
                p["temp"].setValue(round(p_data.get("temp_min", 10)))

                p["use_sun"].setChecked(
                    p_data.get("nocturne_actif", False)
                )

                if "reference_debut" in p_data:
                    p["start_ref"].setCurrentIndex(
                        0 if p_data.get("reference_debut") == "coucher"
                        else 1
                    )
                    p["offset_start"].setValue(
                        p_data.get("decalage_debut", 0)
                    )
                else:
                    p["start_ref"].setCurrentIndex(0)
                    p["offset_start"].setValue(
                        p_data.get("decalage_coucher", 0)
                    )

                if "reference_fin" in p_data:
                    p["end_ref"].setCurrentIndex(
                        0 if p_data.get("reference_fin") == "lever"
                        else 1
                    )
                    p["offset_end"].setValue(
                        p_data.get("decalage_fin", 0)
                    )
                else:
                    p["end_ref"].setCurrentIndex(0)
                    p["offset_end"].setValue(
                        p_data.get("decalage_lever", 0)
                    )

    # ======================================================

    def load_plans_from_path(self, in_path):
        """Charge les plans directement depuis un chemin donné,
        sans boîte de dialogue (utilisé par le chargement global
        du projet)."""

        with open(in_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        self._apply_plans_data(data)

    # ======================================================

    def load_plans_from_file(self):
        """
        Charge une configuration de plans de bridage depuis un
        fichier .json précédemment enregistré, en remplaçant les
        plans actuellement affichés.

        Accepte deux formats :
        - le format natif Bridage (plusieurs plans, clé "plans")
        - le format d'un scénario unique enregistré depuis Albat
          Scenar (clé "periodes" directement à la racine), qui est
          alors chargé comme un plan unique.
        """

        try:

            in_path, _ = QFileDialog.getOpenFileName(
                self,
                "Charger des plans de bridage",
                "",
                "Fichier de plans (*.json)"
            )

            if not in_path:
                return

            self.load_plans_from_path(in_path)

            QMessageBox.information(
                self,
                "Succès",
                f"{len(self.plans)} plan(s) chargé(s) depuis :\n{in_path}"
            )

        except Exception as e:
            QMessageBox.critical(
                self, "Erreur",
                f"Impossible de charger les plans :\n{e}"
            )

    # ======================================================

    def add_period(self, plan):

        frame = QFrame()
        frame.setStyleSheet("""
            QFrame{
                background:rgba(148,116,76,215);
                border:1px solid rgba(255,255,255,45);
                border-radius:16px;
            }
        """)

        layout = QVBoxLayout(frame)
        layout.setContentsMargins(14, 12, 14, 12)
        layout.setSpacing(8)

        line1 = QHBoxLayout()
        line1.setSpacing(4)
        line1.setContentsMargins(0,0,0,0)
        line1.setSpacing(14)
        style_btn = """
        QPushButton{
            background:rgba(120,150,90,180);
            border:1px solid rgba(255,255,255,35);
            border-radius:6px;
            color:white;
            font-size:12px;
            font-weight:900;
            padding:0px;
        }

        QPushButton:hover{
            background:rgba(160,210,120,220);
        }

        QPushButton:pressed{
            background:rgba(90,120,70,255);
        }
        """

        # =====================================
        # STYLE DATE
        # =====================================

        date_style = """
        QDateEdit{
            background:#081008;

            border-radius:14px;

            padding-left:12px;
            padding-right:34px;

            border:1px solid rgba(255,255,255,35);

            color:#eef7d0;

            font-size:15px;
        }

        QDateEdit::drop-down{
            width:1px;
            border:none;
}
        """

        # =====================================
        # START
        # =====================================

        start = QDateEdit()
        start.setCalendarPopup(True)
        start.setDate(QDate.currentDate())
        start.setCalendarPopup(True)
        btn_start = QPushButton("▼")
        btn_start.setFixedSize(18, 18)
        btn_start.setStyleSheet("""
            QPushButton{
                border:none;
                font-size:12px;
                padding:0px;
            }
            QPushButton:hover{
                background:#dddddd;
                border-radius:4px;
            }
        """)

        btn_start.clicked.connect(lambda: self.show_calendar(start))
        start.setDisplayFormat("dd/MM")
        start.setDate(QDate(2025, 4, 1))

        start.setFixedWidth(95)

        start.setStyleSheet(date_style)
       


        # =====================================
        # END
        # =====================================

        end = QDateEdit()
        btn_end = QPushButton("▼")
        btn_end.setFixedSize(18, 18)
        btn_end.setStyleSheet("""
            QPushButton{
                border:none;
                font-size:12px;
                padding:0px;
            }
            QPushButton:hover{
                background:#dddddd;
                border-radius:4px;
            }
        """)

        btn_end.clicked.connect(lambda: self.show_calendar(end))

        end.setStyleSheet(date_style)
        end.setCalendarPopup(True)
        end.setDisplayFormat("dd/MM")
        end.setDate(QDate(2025, 10, 31))
        end.setFixedWidth(95)

        start.setStyleSheet(date_style)
        end.setStyleSheet(date_style)

        delete_btn = QPushButton("✕")
        delete_btn.setFixedSize(20, 20)
        delete_btn.setFont(
            QFont("Arial", 5, QFont.Bold)
        )

        delete_btn.setStyleSheet("""
        QPushButton{
            background:rgba(110,140,80,180);
            border:2px solid rgba(255,255,255,40);
            border-radius:29px;

            color:white;

            font-size:12px;
            font-weight:900;

            padding:0px;

        }

        QPushButton:hover{
            background:rgba(140,180,100,220);
            border:2px solid rgba(220,255,220,120);
        }

        QPushButton:pressed{
            background:rgba(90,120,70,255);
        }
        """)

        shadow = QGraphicsDropShadowEffect()

        shadow.setBlurRadius(25)

        shadow.setXOffset(0)
        shadow.setYOffset(0)

        shadow.setColor(Qt.green)

        delete_btn.setGraphicsEffect(shadow)

        # =====================================
        # LABEL DEBUT
        # =====================================

        lab1 = QLabel("Début")

        lab1.setFixedWidth(50)

        lab1.setAlignment(Qt.AlignCenter)

        lab1.setStyleSheet("""
        font-size:15px;
        font-weight:700;
        background:transparent;
        """)

        line1.addStretch()
        line1.addWidget(lab1)

        # =====================================
        # START
        # =====================================

        start_zone = QHBoxLayout()

        start_zone.setSpacing(6)

        start_zone.addWidget(start)
        start_zone.addWidget(btn_start)

        line1.addLayout(start_zone)

        # =====================================
        # LABEL FIN
        # =====================================

        lab2 = QLabel("Fin")

        lab2.setFixedWidth(40)

        lab2.setAlignment(Qt.AlignCenter)

        lab2.setStyleSheet("""
        font-size:15px;
        font-weight:700;
        background:transparent;
        """)

        line1.addWidget(lab2)

        # =====================================
        # END
        # =====================================

        end_zone = QHBoxLayout()

        end_zone.setSpacing(6)

        end_zone.addWidget(end)
        end_zone.addWidget(btn_end)

        line1.addLayout(end_zone)

        # =====================================
        # BOUTON DELETE
        # =====================================

        line1.addStretch()

        line1.addStretch()
        line1.addWidget(delete_btn)
        
        
        line2 = QHBoxLayout()
        line2.setSpacing(4)
        line2.setContentsMargins(0,0,0,0)


        temp = QSpinBox()
        temp.setRange(-20, 50)
        temp.setValue(10)
        temp.setSuffix(" °C")

        # =========================
        # VENT
        # =========================

        vent = QSpinBox()
        vent.setButtonSymbols(QSpinBox.NoButtons)

        vent.setRange(0, 50)
        vent.setValue(6)
        vent.setSuffix(" m/s")

        vent.setAlignment(Qt.AlignCenter)

        vent_up = QPushButton("▲")
        vent_down = QPushButton("▼")

        vent_up.setFixedSize(20,18)
        vent_down.setFixedSize(20,18)
        vent_up.setStyleSheet(style_btn)
        vent_down.setStyleSheet(style_btn)

        vent_up.clicked.connect(
            lambda: vent.stepUp()
        )

        vent_down.clicked.connect(
            lambda: vent.stepDown()
        )

        vent_box = QVBoxLayout()

        vent_box.setSpacing(2)

        vent_box.addWidget(vent_up)
        vent_box.addWidget(vent_down)

        vent_zone = QHBoxLayout()

        vent_zone.setSpacing(6)

        vent_zone.addWidget(vent)
        vent_zone.addLayout(vent_box)

        lab_vent = QLabel("Vent <")

        lab_vent.setAlignment(Qt.AlignCenter)

        lab_vent.setFixedWidth(65)

        no_vent_limit = QCheckBox("Sans limite")
        no_vent_limit.setStyleSheet("""
            QCheckBox{ background:transparent; spacing:6px; font-size:11px; }
            QCheckBox::indicator{
                width:14px; height:14px;
                border:1px solid rgba(255,255,255,90);
                border-radius:4px; background:rgba(0,0,0,0.35);
            }
            QCheckBox::indicator:hover{ border:1px solid rgba(255,255,255,150); }
            QCheckBox::indicator:checked{
                background:rgba(160,210,120,255);
                border:1px solid rgba(255,255,255,150);
            }
        """)

        def _toggle_vent_limit(checked):
            vent.setEnabled(not checked)
            vent_up.setEnabled(not checked)
            vent_down.setEnabled(not checked)

        no_vent_limit.toggled.connect(_toggle_vent_limit)

        line2.addWidget(lab_vent)
        line2.addLayout(vent_zone)
        line2.addWidget(no_vent_limit)

        # =========================
        # TEMP
        # =========================

        temp = QSpinBox()

        temp.setButtonSymbols(QSpinBox.NoButtons)

        temp.setRange(-20, 50)
        temp.setValue(10)
        temp.setSuffix(" °C")

        temp.setAlignment(Qt.AlignCenter)

        temp_up = QPushButton("▲")
        temp_down = QPushButton("▼")

        temp_up.setFixedSize(20,18)
        temp_down.setFixedSize(20,18)
        temp_up.setStyleSheet(style_btn)
        temp_down.setStyleSheet(style_btn)

        temp_up.clicked.connect(
            lambda: temp.stepUp()
        )

        temp_down.clicked.connect(
            lambda: temp.stepDown()
        )

        temp_box = QVBoxLayout()

        temp_box.setSpacing(2)

        temp_box.addWidget(temp_up)
        temp_box.addWidget(temp_down)

        temp_zone = QHBoxLayout()

        temp_zone.setSpacing(6)

        temp_zone.addWidget(temp)
        temp_zone.addLayout(temp_box)

        lab_temp = QLabel("Temp >")

        lab_temp.setAlignment(Qt.AlignCenter)

        lab_temp.setFixedWidth(65)

        no_temp_limit = QCheckBox("Sans limite")
        no_temp_limit.setStyleSheet("""
            QCheckBox{ background:transparent; spacing:6px; font-size:11px; }
            QCheckBox::indicator{
                width:14px; height:14px;
                border:1px solid rgba(255,255,255,90);
                border-radius:4px; background:rgba(0,0,0,0.35);
            }
            QCheckBox::indicator:hover{ border:1px solid rgba(255,255,255,150); }
            QCheckBox::indicator:checked{
                background:rgba(160,210,120,255);
                border:1px solid rgba(255,255,255,150);
            }
        """)

        def _toggle_temp_limit(checked):
            temp.setEnabled(not checked)
            temp_up.setEnabled(not checked)
            temp_down.setEnabled(not checked)

        no_temp_limit.toggled.connect(_toggle_temp_limit)

        line2.addWidget(lab_temp)
        line2.addLayout(temp_zone)
        line2.addWidget(no_temp_limit)

        layout.addLayout(line1)
        layout.addLayout(line2)

        # =========================
        # RESTRICTION HORAIRE BASÉE SUR LE SOLEIL (propre à cette
        # période) — début et fin réglables indépendamment sur le
        # coucher ou le lever
        # =========================

        use_sun = QCheckBox(
            "Restreindre cette période à une plage basée sur le soleil"
        )
        use_sun.setStyleSheet("""
            QCheckBox{
                background:transparent;
                spacing:8px;
                font-size:11px;
            }
            QCheckBox::indicator{
                width:14px;
                height:14px;
                border:1px solid rgba(255,255,255,90);
                border-radius:4px;
                background:rgba(0,0,0,0.35);
            }
            QCheckBox::indicator:hover{
                border:1px solid rgba(255,255,255,150);
            }
            QCheckBox::indicator:checked{
                background:rgba(160,210,120,255);
                border:1px solid rgba(255,255,255,150);
            }
        """)
        layout.addWidget(use_sun)

        combo_style = """
            QComboBox{
                background:black;
                color:white;
                border:1px solid rgba(255,255,255,45);
                border-radius:8px;
                padding:2px 6px;
                font-size:11px;
            }
        """

        # --- DÉBUT ---
        start_line = QHBoxLayout()

        lab_start = QLabel("Début")
        lab_start.setStyleSheet("background:transparent; font-size:11px;")
        lab_start.setFixedWidth(38)

        start_ref = QComboBox()
        start_ref.addItems(["Coucher", "Lever"])
        start_ref.setStyleSheet(combo_style)
        start_ref.setFixedWidth(90)

        offset_start = QSpinBox()
        offset_start.setButtonSymbols(QSpinBox.NoButtons)
        offset_start.setRange(-600, 600)
        offset_start.setSingleStep(5)
        offset_start.setValue(0)
        offset_start.setSuffix(" min")
        offset_start.setAlignment(Qt.AlignCenter)
        offset_start.setEnabled(False)

        start_up = QPushButton("▲")
        start_down = QPushButton("▼")
        start_up.setFixedSize(20, 18)
        start_down.setFixedSize(20, 18)
        start_up.setStyleSheet(style_btn)
        start_down.setStyleSheet(style_btn)
        start_up.clicked.connect(lambda: offset_start.stepUp())
        start_down.clicked.connect(lambda: offset_start.stepDown())

        start_arrows = QVBoxLayout()
        start_arrows.setSpacing(2)
        start_arrows.addWidget(start_up)
        start_arrows.addWidget(start_down)

        start_line.addWidget(lab_start)
        start_line.addWidget(start_ref)
        start_line.addWidget(offset_start)
        start_line.addLayout(start_arrows)

        layout.addLayout(start_line)

        # --- FIN ---
        end_line = QHBoxLayout()

        lab_end = QLabel("Fin")
        lab_end.setStyleSheet("background:transparent; font-size:11px;")
        lab_end.setFixedWidth(38)

        end_ref = QComboBox()
        end_ref.addItems(["Lever", "Coucher"])
        end_ref.setStyleSheet(combo_style)
        end_ref.setFixedWidth(90)

        offset_end = QSpinBox()
        offset_end.setButtonSymbols(QSpinBox.NoButtons)
        offset_end.setRange(-600, 600)
        offset_end.setSingleStep(5)
        offset_end.setValue(0)
        offset_end.setSuffix(" min")
        offset_end.setAlignment(Qt.AlignCenter)
        offset_end.setEnabled(False)

        end_up = QPushButton("▲")
        end_down = QPushButton("▼")
        end_up.setFixedSize(20, 18)
        end_down.setFixedSize(20, 18)
        end_up.setStyleSheet(style_btn)
        end_down.setStyleSheet(style_btn)
        end_up.clicked.connect(lambda: offset_end.stepUp())
        end_down.clicked.connect(lambda: offset_end.stepDown())

        end_arrows = QVBoxLayout()
        end_arrows.setSpacing(2)
        end_arrows.addWidget(end_up)
        end_arrows.addWidget(end_down)

        end_line.addWidget(lab_end)
        end_line.addWidget(end_ref)
        end_line.addWidget(offset_end)
        end_line.addLayout(end_arrows)

        layout.addLayout(end_line)

        sun_note = QLabel(
            "Décalage négatif = avant, positif = après. "
            "Ex : Début = Coucher, -60 min ; Fin = Coucher, +180 min."
        )
        sun_note.setWordWrap(True)
        sun_note.setStyleSheet(
            "color:#f2e6d0; font-size:9px; background:transparent;"
        )
        layout.addWidget(sun_note)

        def _toggle_sun_limit(checked, _off1=offset_start, _off2=offset_end):
            if checked and not ASTRAL_OK:
                QMessageBox.warning(
                    self, "Module manquant",
                    "Le module 'astral' n'est pas installé.\n"
                    "Exécutez : pip install astral"
                )
                use_sun.setChecked(False)
                return
            _off1.setEnabled(checked)
            _off2.setEnabled(checked)

        use_sun.toggled.connect(_toggle_sun_limit)

        plan["period_container"].addWidget(frame)

        period_data = {
            "frame": frame,
            "start": start,
            "end": end,
            "vent": vent,
            "temp": temp,
            "no_vent_limit": no_vent_limit,
            "no_temp_limit": no_temp_limit,
            "use_sun": use_sun,
            "start_ref": start_ref,
            "offset_start": offset_start,
            "end_ref": end_ref,
            "offset_end": offset_end,
        }

        plan["periods"].append(period_data)

        delete_btn.clicked.connect(
            lambda _, p=plan, pd=period_data:
            self.remove_period(p, pd)
        )

    # ======================================================
    

    def remove_period(self, plan, period_data):

        frame = period_data["frame"]

        plan["period_container"].removeWidget(frame)

        frame.setParent(None)

        frame.deleteLater()

        if period_data in plan["periods"]:
            plan["periods"].remove(period_data)

    # ======================================================

    # ======================================================

    def _lookup_city_geoapi(self, name):
        """
        Recherche une commune française via l'API officielle
        geo.api.gouv.fr. Contrairement à la base hors ligne
        d'astral (qui ne connaît que quelques grandes villes dans
        le monde), cette API couvre les ~35000 communes de France,
        y compris les petites communes du Grand Est.

        Retourne un objet LocationInfo (astral) ou None si la
        ville est introuvable / pas de connexion internet.
        """

        import urllib.request
        import urllib.parse
        import json as _json

        try:
            query = urllib.parse.urlencode({
                "nom": name,
                "fields": "nom,centre,population,codeDepartement",
                "boost": "population",
                "limit": 1,
            })
            url = f"https://geo.api.gouv.fr/communes?{query}"

            req = urllib.request.Request(
                url, headers={"User-Agent": "AlbatBridage/1.0"}
            )

            with urllib.request.urlopen(req, timeout=6) as response:
                data = _json.loads(response.read().decode("utf-8"))

            if not data:
                return None

            commune = data[0]
            lon, lat = commune["centre"]["coordinates"]

            return LocationInfo(
                name=commune["nom"],
                region=f"France (dép. {commune.get('codeDepartement', '?')})",
                timezone="Europe/Paris",
                latitude=lat,
                longitude=lon,
            )

        except Exception:
            return None

    # ======================================================

    def search_city(self):

        name = self.city_input.text().strip()

        if not name:
            self.city_status.setText("Veuillez saisir un nom de ville.")
            return

        info = None

        try:
            info = astral_lookup(name, astral_database())

            if isinstance(info, dict):
                # Un groupe régional a été trouvé, pas une ville précise
                info = None

        except KeyError:
            info = None

        if info is None:
            # Pas trouvée dans la petite base hors ligne d'astral :
            # on interroge la base officielle des communes françaises
            # (couvre toutes les communes, quelle que soit leur taille).
            self.city_status.setText("Recherche en ligne...")
            QApplication.processEvents()
            info = self._lookup_city_geoapi(name)

        if info is not None:

            self.location_info = info

            self.city_status.setText(
                f"Trouvée : {info.name} / {info.region} "
                f"(lat {info.latitude:.2f}, lon {info.longitude:.2f}) "
                f"— fuseau {info.timezone}"
            )

        else:

            self.location_info = None

            self.city_status.setText(
                "Ville introuvable. Vérifiez l'orthographe, ou votre "
                "connexion internet (la recherche interroge la base "
                "officielle des communes de France)."
            )

    # ======================================================

    def get_sun_window(self, date, sun_cache):
        """
        Retourne (coucher, lever_lendemain) en UTC+2 fixe (naive),
        cohérent avec la convention SCADA du projet.
        On convertit via UTC puis on ajoute +2h fixes, sans jamais
        utiliser .replace(tzinfo=None) directement (ce qui donnait
        UTC+1 après le changement d'heure du 26 octobre).
        """

        if date in sun_cache:
            return sun_cache[date]

        from zoneinfo import ZoneInfo as _ZI

        tz  = _ZI(self.location_info.timezone)
        UTC = _ZI("UTC")

        s_today = astral_sun(
            self.location_info.observer,
            date=date,
            tzinfo=tz
        )

        s_tomorrow = astral_sun(
            self.location_info.observer,
            date=date + dtmod.timedelta(days=1),
            tzinfo=tz
        )

        def to_utc2(aware_dt):
            return (
                aware_dt.astimezone(UTC).replace(tzinfo=None)
                + dtmod.timedelta(hours=2)
            )

        result = (
            to_utc2(s_today["sunset"]),
            to_utc2(s_tomorrow["sunrise"])
        )

        sun_cache[date] = result

        return result

    # ======================================================

    def pick_file(self):

        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Choisir les fichiers Excel",
            "",
            "Excel (*.xlsx *.xlsm *.xls)"
        )

        if files:

            self.files = files

            if len(files) == 1:

                label = os.path.basename(files[0])

            else:

                label = f"{len(files)} fichiers sélectionnés"

            self.lbl_file.setText(label)

    # ======================================================

    def check_bridage(self, plan, dt, vent_value, temp_value, sun_cache=None):

        for p in plan["periods"]:

            s = p["start"].date()
            e = p["end"].date()

            start_val = s.month() * 100 + s.day()
            end_val = e.month() * 100 + e.day()

            current_val = dt.month * 100 + dt.day

            in_period = (
                start_val <= current_val <= end_val
            )

            vent_ok = (
                p["no_vent_limit"].isChecked()
                or vent_value < p["vent"].value()
            )

            temp_ok = (
                p["no_temp_limit"].isChecked()
                or temp_value > p["temp"].value()
            )

            if not (in_period and vent_ok and temp_ok):
                continue

            # ------------------------------------------------
            # Restriction horaire basée sur le soleil, si activée
            # (début et fin réglables indépendamment sur le
            # coucher ou le lever)
            # ------------------------------------------------
            if p["use_sun"].isChecked():

                if self.location_info is None:
                    # Pas de ville valide renseignée : impossible
                    # de vérifier la condition
                    continue

                ref_date = dt.date()

                if dt.time() < dtmod.time(12, 0):
                    ref_date = ref_date - dtmod.timedelta(days=1)

                sunset_t, sunrise_t = self.get_sun_window(
                    ref_date, sun_cache
                )

                start_base = (
                    sunset_t if p["start_ref"].currentIndex() == 0
                    else sunrise_t
                )
                end_base = (
                    sunrise_t if p["end_ref"].currentIndex() == 0
                    else sunset_t
                )

                window_start = start_base + dtmod.timedelta(
                    minutes=p["offset_start"].value()
                )

                window_end = end_base + dtmod.timedelta(
                    minutes=p["offset_end"].value()
                )

                if not (window_start <= dt <= window_end):
                    continue

            return True

        return False

    # ======================================================

    def _style_worksheet(self, ws):
        """
        Applique un style homogène à une feuille de données :
        en-tête vert avec texte blanc, colonnes dimensionnées
        selon leur contenu, formats de date/heure corrects,
        police cohérente, et lignes figées sous l'en-tête.
        """

        header_font = Font(name="Arial", bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4C6B3A")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )

        for col_idx, col_cells in enumerate(
            ws.iter_cols(min_row=1, max_row=ws.max_row), start=1
        ):
            header_text = str(col_cells[0].value or "")
            max_len = len(header_text)
            sample_val = None

            for cell in col_cells[1:]:

                if cell.value is not None:

                    if sample_val is None:
                        sample_val = cell.value

                    length = len(str(cell.value))

                    if length > max_len:
                        max_len = length

                cell.font = Font(name="Arial")
                cell.alignment = Alignment(horizontal="center")

            width = min(max(max_len + 3, 10), 34)
            ws.column_dimensions[
                get_column_letter(col_idx)
            ].width = width

            if isinstance(sample_val, dtmod.datetime):

                is_full_datetime = any(
                    c.value is not None
                    and isinstance(c.value, dtmod.datetime)
                    and c.value.time() != dtmod.time(0, 0)
                    for c in col_cells[1:200]
                )

                fmt = (
                    "dd/mm/yyyy hh:mm:ss"
                    if is_full_datetime
                    else "dd/mm/yyyy"
                )

                for cell in col_cells[1:]:
                    if cell.value is not None:
                        cell.number_format = fmt

            elif isinstance(sample_val, dtmod.date):

                for cell in col_cells[1:]:
                    if cell.value is not None:
                        cell.number_format = "dd/mm/yyyy"

        ws.freeze_panes = "A2"

    # ======================================================

    def _build_synthese_sheet(self, wb):
        """
        Ajoute un onglet 'Synthèse des plans' récapitulant, pour
        chaque plan de bridage exécuté : ses périodes (dates,
        seuils vent/température) et son réglage de restriction
        nocturne (décalages coucher/lever).
        """

        ws = wb.create_sheet("Synthèse des plans")

        title_font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
        title_fill = PatternFill("solid", fgColor="4C6B3A")

        plan_font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        plan_fill = PatternFill("solid", fgColor="6E8F4E")

        sub_font = Font(name="Arial", italic=True, size=10, color="444444")

        table_header_font = Font(name="Arial", bold=True, color="FFFFFF")
        table_header_fill = PatternFill("solid", fgColor="8FAE6B")

        normal_font = Font(name="Arial", size=10)

        row = 1

        ws.cell(row=row, column=1, value="SYNTHÈSE DES PLANS DE BRIDAGE")
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=4
        )
        cell = ws.cell(row=row, column=1)
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

        info_year = self.ref_year if self.ref_year else dtmod.date.today().year
        info_park = self.park_name if self.park_name else "-"

        ws.cell(
            row=row, column=1,
            value=f"Parc éolien : {info_park}    |    Année : {info_year}"
        )
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=4
        )
        cell = ws.cell(row=row, column=1)
        cell.font = Font(name="Arial", italic=True, size=10, color="444444")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 2

        for plan in self.plans:

            plan_name = plan["name_edit"].text().strip() or "Plan sans nom"

            ws.cell(row=row, column=1, value=plan_name)
            ws.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=4
            )
            cell = ws.cell(row=row, column=1)
            cell.font = plan_font
            cell.fill = plan_fill
            cell.alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            ws.row_dimensions[row].height = 20
            row += 1

            headers = ["Période", "Vent <", "Temp >", "Restriction nocturne"]
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=c, value=h)
                cell.font = table_header_font
                cell.fill = table_header_fill
                cell.alignment = Alignment(horizontal="center")
            row += 1

            for p in plan["periods"]:

                start_txt = p["start"].date().toString("dd/MM")
                end_txt = p["end"].date().toString("dd/MM")

                vent_txt = (
                    "Sans limite"
                    if p["no_vent_limit"].isChecked()
                    else f"{p['vent'].value()} m/s"
                )

                temp_txt = (
                    "Sans limite"
                    if p["no_temp_limit"].isChecked()
                    else f"{p['temp'].value()} °C"
                )

                if p["use_sun"].isChecked():
                    sun_txt = (
                        f"Début: {p['start_ref'].currentText()} "
                        f"{p['offset_start'].value():+d}min → "
                        f"Fin: {p['end_ref'].currentText()} "
                        f"{p['offset_end'].value():+d}min "
                        f"(ville : {self.city_input.text().strip() or '?'})"
                    )
                else:
                    sun_txt = "Aucune (jour et nuit)"

                ws.cell(
                    row=row, column=1,
                    value=f"{start_txt} → {end_txt}"
                ).font = normal_font

                ws.cell(
                    row=row, column=2, value=vent_txt
                ).font = normal_font
                ws.cell(row=row, column=2).alignment = Alignment(
                    horizontal="center"
                )

                ws.cell(
                    row=row, column=3, value=temp_txt
                ).font = normal_font
                ws.cell(row=row, column=3).alignment = Alignment(
                    horizontal="center"
                )

                ws.cell(
                    row=row, column=4, value=sun_txt
                ).font = normal_font
                ws.cell(row=row, column=4).alignment = Alignment(
                    horizontal="center"
                )

                row += 1

            row += 1

        for col_letter, width in zip("ABCD", [22, 14, 14, 46]):
            ws.column_dimensions[col_letter].width = width

    # ======================================================

    def run(self):

        try:

            if not self.files:
                raise Exception(
                    "Sélectionnez au moins un fichier."
                )

            if not self.plans:
                raise Exception(
                    "Ajoutez au moins un plan de bridage."
                )

            any_plan_uses_sun = any(
                p["use_sun"].isChecked()
                for plan in self.plans
                for p in plan["periods"]
            )

            if any_plan_uses_sun and self.location_info is None:
                raise Exception(
                    "Recherchez et validez une ville de référence "
                    "avant de lancer le bridage nocturne."
                )

            sun_cache = {}

            single_file = (len(self.files) == 1)

            if single_file:

                out_year = self.ref_year if self.ref_year else dtmod.date.today().year

                park_part = self.park_name if self.park_name else "Albat"

                default_name = f"{park_part}_{out_year}_Bridage.xlsx"

                if self.default_save_dir and os.path.isdir(self.default_save_dir):
                    default_name = os.path.join(
                        self.default_save_dir, default_name
                    )

                out_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Enregistrer le fichier de sortie",
                    default_name,
                    "Excel (*.xlsx)"
                )

                if not out_path:
                    return

                if not out_path.lower().endswith(".xlsx"):
                    out_path += ".xlsx"

                out_dir = os.path.dirname(out_path)

            else:

                out_dir = QFileDialog.getExistingDirectory(
                    self,
                    "Choisir le dossier de sortie",
                    self.default_save_dir or ""
                )

                if not out_dir:
                    return

                suffix, ok = QInputDialog.getText(
                    self,
                    "Nom des fichiers",
                    "Suffixe ajouté au nom de chaque fichier d'origine :",
                    text="_BRIDAGE"
                )

                if not ok:
                    return

                suffix = suffix.strip() or "_BRIDAGE"

            total_files = len(self.files)

            for file_index, current_file in enumerate(self.files):

                progress_base = int((file_index / total_files) * 100)

                self.pb.setValue(progress_base + 5)

                df = pd.read_excel(current_file)

                cols = {
                    c.lower().strip(): c
                    for c in df.columns
                }

                if "date" not in cols:
                    raise Exception(
                        f"Colonne date introuvable dans : {os.path.basename(current_file)}"
                    )

                if "vent" not in cols:
                    raise Exception(
                        f"Colonne vent introuvable dans : {os.path.basename(current_file)}"
                    )

                temp_col = None

                for c in cols:

                    if "temp" in c:
                        temp_col = cols[c]

                if not temp_col:
                    raise Exception(
                        f"Colonne température introuvable dans : {os.path.basename(current_file)}"
                    )

                # Pré-parsing des lignes (date / vent / temp), une seule
                # fois, puis on applique chaque plan dessus.
                #
                # Priorité pour la date/heure exacte :
                #   1. colonne 'date_heure' si elle existe (datetime complet)
                #   2. sinon combinaison de 'date' + 'heure'
                #   3. sinon 'date' seul (heure perdue → risque de décalage)

                if "date_heure" in cols:
                    date_col_key = "date_heure"
                    heure_col_key = None
                elif "date" in cols and "heure" in cols:
                    date_col_key = "date"
                    heure_col_key = "heure"
                else:
                    date_col_key = "date"
                    heure_col_key = None

                parsed_rows = []

                for _, row in df.iterrows():

                    try:

                        if heure_col_key:
                            date_part = pd.to_datetime(row[cols[date_col_key]]).date()
                            heure_str = str(row[cols[heure_col_key]])
                            dt = pd.to_datetime(f"{date_part} {heure_str}")
                        else:
                            dt = pd.to_datetime(row[cols[date_col_key]])

                        vent = float(row[cols["vent"]])
                        temp = float(row[temp_col])
                        parsed_rows.append((dt, vent, temp))

                    except Exception:
                        parsed_rows.append(None)

                summary_rows = []
                used_col_names = set()

                # ------------------------------------------------
                # Colonnes de diagnostic : coucher / lever réellement
                # calculés pour chaque contact (utile pour vérifier
                # les décalages appliqués par les plans).
                # ------------------------------------------------
                any_plan_uses_sun_diag = any(
                    p["use_sun"].isChecked()
                    for plan in self.plans
                    for p in plan["periods"]
                )

                if any_plan_uses_sun_diag and self.location_info is not None:

                    coucher_vals = []
                    lever_vals = []

                    for parsed in parsed_rows:

                        if parsed is None:
                            coucher_vals.append(None)
                            lever_vals.append(None)
                            continue

                        dt_row, _, _ = parsed

                        ref_date = dt_row.date()

                        if dt_row.time() < dtmod.time(12, 0):
                            ref_date = ref_date - dtmod.timedelta(days=1)

                        try:
                            sunset_t, sunrise_t = self.get_sun_window(
                                ref_date, sun_cache
                            )
                            coucher_vals.append(sunset_t)
                            lever_vals.append(sunrise_t)
                        except Exception:
                            coucher_vals.append(None)
                            lever_vals.append(None)

                    df["Coucher du soleil (calculé)"] = coucher_vals
                    df["Lever du soleil (calculé, lendemain)"] = lever_vals


                # ------------------------------------------------
                # Colonnes de diagnostic : coucher / lever réellement
                # calculés pour chaque contact (utile pour vérifier
                # les décalages appliqués par les plans).
                # ------------------------------------------------
                any_plan_uses_sun = any(
                    p["use_sun"].isChecked()
                    for plan in self.plans
                    for p in plan["periods"]
                )

                if any_plan_uses_sun and self.location_info is not None:

                    coucher_vals = []
                    lever_vals = []

                    for parsed in parsed_rows:

                        if parsed is None:
                            coucher_vals.append(None)
                            lever_vals.append(None)
                            continue

                        dt_row, _, _ = parsed

                        ref_date = dt_row.date()

                        if dt_row.time() < dtmod.time(12, 0):
                            ref_date = ref_date - dtmod.timedelta(days=1)

                        try:
                            sunset_t, sunrise_t = self.get_sun_window(
                                ref_date, sun_cache
                            )
                            coucher_vals.append(sunset_t)
                            lever_vals.append(sunrise_t)
                        except Exception:
                            coucher_vals.append(None)
                            lever_vals.append(None)

                    df["Coucher du soleil (calculé)"] = coucher_vals
                    df["Lever du soleil (calculé, lendemain)"] = lever_vals

                for plan_index, plan in enumerate(self.plans):

                    plan_name = plan["name_edit"].text().strip()

                    if plan_name == "":
                        plan_name = f"Plan {plan_index + 1}"

                    # Évite les collisions de noms de colonnes si deux
                    # plans portent le même nom
                    bridage_col = plan_name
                    suffix = 2

                    while bridage_col in used_col_names or bridage_col in df.columns:
                        bridage_col = f"{plan_name} ({suffix})"
                        suffix += 1

                    used_col_names.add(bridage_col)

                    results = []

                    for parsed in parsed_rows:

                        if parsed is None:
                            results.append("Non")
                            continue

                        dt, vent, temp = parsed

                        try:
                            bridage = self.check_bridage(
                                plan,
                                dt,
                                vent,
                                temp,
                                sun_cache
                            )

                            results.append(
                                "Oui" if bridage
                                else "Non"
                            )

                        except Exception:
                            results.append("Non")

                    df[bridage_col] = results

                    yes_count = results.count("Oui")
                    no_count = results.count("Non")
                    total_count = yes_count + no_count

                    protection_pct = (
                        round(yes_count / total_count * 100, 2)
                        if total_count > 0 else 0
                    )

                    summary_rows.append({
                        "Plan": bridage_col,
                        "Contacts non concernés par le bridage": no_count,
                        "Contacts concernés par le bridage": yes_count,
                        "Total des contacts": total_count,
                        "Taux de protection (%)": protection_pct,
                    })

                resume = pd.DataFrame(summary_rows)

                if single_file:
                    out = out_path
                else:
                    original_name = os.path.splitext(
                        os.path.basename(current_file)
                    )[0]

                    out = os.path.join(
                        out_dir,
                        f"{original_name}{suffix}.xlsx"
                    )

                with pd.ExcelWriter(out, engine="openpyxl") as writer:

                    df.to_excel(
                        writer,
                        sheet_name="Données",
                        index=False
                    )

                    resume.to_excel(
                        writer,
                        sheet_name="Résumé",
                        index=False
                    )

                wb_out = openpyxl.load_workbook(out)

                self._style_worksheet(wb_out["Données"])
                self._style_worksheet(wb_out["Résumé"])

                self._build_synthese_sheet(wb_out)

                wb_out.save(out)

                progress = int(((file_index + 1) / total_files) * 100)

                self.pb.setValue(progress)

            QMessageBox.information(
                self,
                "Succès",
                f"{total_files} fichier(s) traité(s) avec succès."
            )

        except Exception as e:

            QMessageBox.critical(
                self,
                "Erreur",
                str(e)
            )
# ==========================================================
# MAIN
# ==========================================================

if __name__ == "__main__":

    app = QApplication(sys.argv)

    try:
        app.setStyle("Fusion")
    except:
        pass

    win = BridageWindow()
    win.show()

    sys.exit(app.exec())