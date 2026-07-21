
# ==========================================================
# ALBAT SCENAR - ONGLETS COMPLETS + UI RESTAUREE
# ==========================================================

import sys
import os
import pandas as pd
import datetime as dtmod

try:
    from astral import LocationInfo
    from astral.sun import sun as astral_sun
    from zoneinfo import ZoneInfo
    ASTRAL_OK = True
except ImportError:
    ASTRAL_OK = False

from PySide6.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton,
    QFileDialog, QVBoxLayout, QMessageBox,
    QSlider, QTableWidget, QTableWidgetItem,
    QSpinBox, QDoubleSpinBox,
    QHeaderView, QLineEdit, QDateEdit,
    QHBoxLayout, QCheckBox, QTabWidget,
    QInputDialog, QComboBox
)

from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QPixmap
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


from modules.utils import resource_path

BG = resource_path(
    "assets",
    "background_graph.png"
)


# ==========================================================
# ONGLET PERIODE
# ==========================================================

class PeriodTab(QWidget):

    def __init__(self, parent_window):
        super().__init__()

        self.parent_window = parent_window

        layout = QVBoxLayout(self)

        layout.setSpacing(8)

        # ==========================================
        # DATES
        # ==========================================

        dates = QHBoxLayout()

        self.date_start = QDateEdit()
        self.date_start.setCalendarPopup(True)
        self.date_start.setDate(QDate(2025, 4, 1))

        self.date_end = QDateEdit()
        self.date_end.setCalendarPopup(True)
        self.date_end.setDate(QDate(2025, 10, 31))

        dates.addWidget(QLabel("Date début"))
        dates.addWidget(self.date_start)

        dates.addWidget(QLabel("Date fin"))
        dates.addWidget(self.date_end)

        layout.addLayout(dates)

        # ==========================================
        # TAUX DE PROTECTION GLOBAL (mis à jour en direct)
        # ==========================================

        self.lab_global_pct = QLabel("Taux de protection de la plage : —")

        self.lab_global_pct.setAlignment(Qt.AlignCenter)

        self.lab_global_pct.setStyleSheet("""
            background: rgba(90,127,71,190);
            color: white;
            font-size: 18px;
            font-weight: 900;
            border-radius: 12px;
            padding: 8px;
            border: 1px solid rgba(255,255,255,45);
        """)

        layout.addWidget(self.lab_global_pct)

        # ==========================================
        # TABLEAU
        # ==========================================

        self.table = QTableWidget()

        self.table.setColumnCount(4)

        self.table.setHorizontalHeaderLabels([
            "Espèce",
            "Protégés",
            "Total",
            "% Protection"
        ])

        self.table.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )

        self.table.setMinimumHeight(420)

        self.table.verticalHeader().setDefaultSectionSize(30)

        layout.addWidget(self.table, stretch=3)

        layout.addStretch(0)

        # ==========================================
        # VENT
        # ==========================================

        self.check_vent = QCheckBox(
            "Activer bridage vent"
        )

        self.check_vent.setChecked(True)

        vent_box = QVBoxLayout()

        vent_box.setSpacing(2)

        self.check_vent.setMaximumHeight(22)

        vent_box.addWidget(self.check_vent)

        self.lab_vent = QLabel(
            "Vent ≤ 5.0 m/s"
        )

        self.lab_vent.setAlignment(Qt.AlignCenter)

        self.lab_vent.setMaximumHeight(18)

        vent_box.addWidget(self.lab_vent)

        vent_row = QHBoxLayout()
        vent_row.setSpacing(6)

        self.slider_vent = QSlider(Qt.Horizontal)

        self.slider_vent.setRange(0, 150)
        self.slider_vent.setValue(50)

        self.slider_vent.setMaximumHeight(18)

        self.spin_vent = QDoubleSpinBox()
        self.spin_vent.setRange(0.0, 15.0)
        self.spin_vent.setSingleStep(0.1)
        self.spin_vent.setDecimals(1)
        self.spin_vent.setSuffix(" m/s")
        self.spin_vent.setValue(5.0)
        self.spin_vent.setFixedWidth(90)

        self.slider_vent.valueChanged.connect(
            lambda v: self.spin_vent.setValue(v / 10)
            if abs(self.spin_vent.value() - v / 10) > 0.001 else None
        )
        self.spin_vent.valueChanged.connect(
            lambda v: self.slider_vent.setValue(round(v * 10))
            if self.slider_vent.value() != round(v * 10) else None
        )

        vent_row.addWidget(self.slider_vent)
        vent_row.addWidget(self.spin_vent)

        vent_box.addLayout(vent_row)

        layout.addLayout(vent_box)

        # ==========================================
        # TEMP
        # ==========================================

        self.check_temp = QCheckBox(
            "Activer bridage température"
        )

        self.check_temp.setChecked(True)

        temp_box = QVBoxLayout()

        temp_box.setSpacing(2)

        self.check_temp.setMaximumHeight(22)

        temp_box.addWidget(self.check_temp)

        self.lab_temp = QLabel(
            "Température ≥ 10.0 °C"
        )

        self.lab_temp.setAlignment(Qt.AlignCenter)

        self.lab_temp.setMaximumHeight(18)

        temp_box.addWidget(self.lab_temp)

        temp_row = QHBoxLayout()
        temp_row.setSpacing(6)

        self.slider_temp = QSlider(Qt.Horizontal)

        self.slider_temp.setRange(-50, 400)
        self.slider_temp.setValue(100)

        self.slider_temp.setMaximumHeight(18)

        self.spin_temp = QDoubleSpinBox()
        self.spin_temp.setRange(-5.0, 40.0)
        self.spin_temp.setSingleStep(0.1)
        self.spin_temp.setDecimals(1)
        self.spin_temp.setSuffix(" °C")
        self.spin_temp.setValue(10.0)
        self.spin_temp.setFixedWidth(90)

        self.slider_temp.valueChanged.connect(
            lambda v: self.spin_temp.setValue(v / 10)
            if abs(self.spin_temp.value() - v / 10) > 0.001 else None
        )
        self.spin_temp.valueChanged.connect(
            lambda v: self.slider_temp.setValue(round(v * 10))
            if self.slider_temp.value() != round(v * 10) else None
        )

        temp_row.addWidget(self.slider_temp)
        temp_row.addWidget(self.spin_temp)

        temp_box.addLayout(temp_row)

        layout.addLayout(temp_box)

        # ==========================================
        # PLAGE HORAIRE BASÉE SUR LE SOLEIL
        # (début et fin réglables indépendamment sur
        # le coucher ou le lever)
        # ==========================================

        self.check_sun = QCheckBox(
            "Restreindre à une plage horaire basée sur le soleil"
        )

        self.check_sun.setChecked(False)

        self.check_sun.setMaximumHeight(22)

        self.check_sun.toggled.connect(self._check_sun_available)

        layout.addWidget(self.check_sun)

        sun_box = QVBoxLayout()
        sun_box.setSpacing(4)

        # --- DÉBUT ---
        start_row = QHBoxLayout()
        start_row.setSpacing(6)

        lab_start = QLabel("Début :")
        lab_start.setStyleSheet("background:transparent; font-size:11px;")
        lab_start.setFixedWidth(42)

        self.start_ref = QComboBox()
        self.start_ref.addItems(["Coucher du soleil", "Lever du soleil"])
        self.start_ref.setFixedWidth(140)

        self.slider_start_offset = QSlider(Qt.Horizontal)
        self.slider_start_offset.setRange(-600, 600)
        self.slider_start_offset.setValue(0)
        self.slider_start_offset.setMaximumHeight(18)

        self.spin_start_offset = QSpinBox()
        self.spin_start_offset.setRange(-600, 600)
        self.spin_start_offset.setSingleStep(1)
        self.spin_start_offset.setSuffix(" min")
        self.spin_start_offset.setValue(0)
        self.spin_start_offset.setFixedWidth(80)

        self.slider_start_offset.valueChanged.connect(
            lambda v: self.spin_start_offset.setValue(v)
            if self.spin_start_offset.value() != v else None
        )
        self.spin_start_offset.valueChanged.connect(
            lambda v: self.slider_start_offset.setValue(v)
            if self.slider_start_offset.value() != v else None
        )

        start_row.addWidget(lab_start)
        start_row.addWidget(self.start_ref)
        start_row.addWidget(self.slider_start_offset)
        start_row.addWidget(self.spin_start_offset)

        sun_box.addLayout(start_row)

        # --- FIN ---
        end_row = QHBoxLayout()
        end_row.setSpacing(6)

        lab_end = QLabel("Fin :")
        lab_end.setStyleSheet("background:transparent; font-size:11px;")
        lab_end.setFixedWidth(42)

        self.end_ref = QComboBox()
        self.end_ref.addItems(["Lever du soleil", "Coucher du soleil"])
        self.end_ref.setFixedWidth(140)

        self.slider_end_offset = QSlider(Qt.Horizontal)
        self.slider_end_offset.setRange(-600, 600)
        self.slider_end_offset.setValue(0)
        self.slider_end_offset.setMaximumHeight(18)

        self.spin_end_offset = QSpinBox()
        self.spin_end_offset.setRange(-600, 600)
        self.spin_end_offset.setSingleStep(1)
        self.spin_end_offset.setSuffix(" min")
        self.spin_end_offset.setValue(0)
        self.spin_end_offset.setFixedWidth(80)

        self.slider_end_offset.valueChanged.connect(
            lambda v: self.spin_end_offset.setValue(v)
            if self.spin_end_offset.value() != v else None
        )
        self.spin_end_offset.valueChanged.connect(
            lambda v: self.slider_end_offset.setValue(v)
            if self.slider_end_offset.value() != v else None
        )

        end_row.addWidget(lab_end)
        end_row.addWidget(self.end_ref)
        end_row.addWidget(self.slider_end_offset)
        end_row.addWidget(self.spin_end_offset)

        sun_box.addLayout(end_row)

        sun_note = QLabel(
            "Décalage négatif = avant l'évènement, positif = après. "
            "Ex : Début = Coucher, -60 min (1h avant le coucher) ; "
            "Fin = Coucher, +180 min (3h après le coucher)."
        )
        sun_note.setWordWrap(True)
        sun_note.setStyleSheet(
            "color:#cfe0a0; font-size:9px; background:transparent;"
        )
        sun_box.addWidget(sun_note)

        layout.addLayout(sun_box)

        # ==========================================
        # SIGNALS
        # ==========================================

        for sig in [
            self.slider_vent.valueChanged,
            self.slider_temp.valueChanged,
            self.check_vent.stateChanged,
            self.check_temp.stateChanged,
            self.date_start.dateChanged,
            self.date_end.dateChanged,
            self.slider_start_offset.valueChanged,
            self.slider_end_offset.valueChanged,
            self.start_ref.currentIndexChanged,
            self.end_ref.currentIndexChanged,
            self.check_sun.stateChanged,
        ]:
            sig.connect(self.update_tab)

    # ======================================================

    # ======================================================

    def _check_sun_available(self, checked):

        if not checked:
            return

        parent = self.parent_window

        if not ASTRAL_OK:
            QMessageBox.warning(
                self, "Module manquant",
                "Le module 'astral' n'est pas installé.\n"
                "Exécutez : pip install astral"
            )
            self.check_sun.setChecked(False)
            return

        if parent.location_info is None:
            QMessageBox.warning(
                self, "Ville non définie",
                "Aucune ville de référence n'a pu être résolue.\n"
                "Vérifiez la ville saisie au démarrage de l'application."
            )
            self.check_sun.setChecked(False)

    # ======================================================

    def compute_mask(self, df, date_col=None):

        vent_limit = self.slider_vent.value() / 10
        temp_limit = self.slider_temp.value() / 10

        vent_active = self.check_vent.isChecked()
        temp_active = self.check_temp.isChecked()

        vent_cond = False
        temp_cond = False

        if vent_active:
            vent_cond = (
                df["vent"] <= vent_limit
            )

        if temp_active:
            temp_cond = (
                df["temp"] >= temp_limit
            )

        if vent_active and temp_active:
            base_mask = vent_cond & temp_cond

        elif vent_active:
            base_mask = vent_cond

        elif temp_active:
            base_mask = temp_cond

        else:
            # Aucune condition vent/température activée : pas de
            # restriction sur ce critère (équivalent "sans limite"
            # dans Bridage) — le bridage dépend alors uniquement
            # de la plage horaire nocturne si elle est activée.
            base_mask = pd.Series(
                [True] * len(df),
                index=df.index
            )

        # ==========================================
        # PLAGE HORAIRE (coucher / lever du soleil)
        # ==========================================

        if self.check_sun.isChecked():

            start_ref_key = (
                "coucher" if self.start_ref.currentIndex() == 0
                else "lever"
            )
            end_ref_key = (
                "lever" if self.end_ref.currentIndex() == 0
                else "coucher"
            )

            night_mask = self.parent_window.compute_night_mask(
                df,
                date_col,
                start_ref_key,
                self.slider_start_offset.value(),
                end_ref_key,
                self.slider_end_offset.value()
            )

            return base_mask & night_mask

        return base_mask

    # ======================================================

    def update_tab(self):

        self.lab_vent.setText(
            f"Vent ≤ {self.slider_vent.value()/10:.1f} m/s"
        )

        self.lab_temp.setText(
            f"Température ≥ {self.slider_temp.value()/10:.1f} °C"
        )

        self.parent_window.update_all_tabs()


# ==========================================================
# FENETRE PRINCIPALE
# ==========================================================

class OptimisationWindow(QWidget):

    def __init__(self, park_name="", ref_year=None, ref_city="", bureau_etude="", default_save_dir="", correlations_dir_hint=""):
        super().__init__()

        self.park_name = park_name
        self.ref_year = ref_year
        self.ref_city = ref_city
        self.bureau_etude = bureau_etude
        self.default_save_dir = default_save_dir
        self.correlations_dir = correlations_dir_hint
        self.df = pd.DataFrame()

        self.location_info = None
        self._sun_cache = {}

        if self.ref_city and ASTRAL_OK:
            self._resolve_ref_city()

        self.setWindowTitle("Albat Scenar")

        self.resize(500, 900)

        self.setMinimumSize(500, 900)

        self.setStyleSheet("""

        QWidget{
            background:#071107;
            color:white;
            font-family:"Segoe UI","Helvetica Neue",Arial,sans-serif;
        }

        QPushButton{
            background:rgba(90,127,71,180);
            border-radius:14px;
            padding:8px;
            font-size:13px;
            font-weight:700;
            color:white;
        }

        QPushButton:hover{
            background:#6b9654;
        }

        QLabel{
            font-size:14px;
        }

        QSlider::groove:horizontal{
            height:6px;
            background:#222;
            border-radius:5px;
        }

        QSlider::handle:horizontal{
            background:#a6f56f;
            width:14px;
            margin:-5px 0;
            border-radius:10px;
        }

        QTableWidget{
            background:rgba(0,0,0,110);
            border-radius:14px;
            color:white;
            gridline-color:rgba(255,255,255,30);
            font-size:14px;
        }

        QHeaderView::section{
            background:#5a7f47;
            color:white;
            padding:8px;
            border:none;
            font-weight:700;
        }

        QTabWidget::pane{
            border:none;
        }

        QTabBar::tab{
            background:rgba(0,0,0,120);
            padding:10px;
            margin:2px;
            border-radius:10px;
        }

        QTabBar::tab:selected{
            background:#5a7f47;
        }

        QLineEdit{
            background:rgba(0,0,0,120);
            border-radius:12px;
            padding:10px;
            color:white;
        }

        """)

        self.build_ui()

        # Auto-chargement du fichier Excel du dossier "Correlations"
        # du projet, pour ne pas avoir à le sélectionner à nouveau
        # manuellement dans chaque onglet. Reprend exactement ce
        # que fait load_excel() (lecture + rafraîchissement), sans
        # passer par le dialogue de sélection.
        auto_file = self._find_correlations_excel()
        if auto_file:
            try:
                self.df = pd.read_excel(auto_file)
                self.file_label.setText(os.path.basename(auto_file))
                self.update_all_tabs()
            except Exception:
                pass

    # ======================================================

    def _find_correlations_excel(self):
        """
        Cherche, dans le dossier 'Correlations' du projet, un
        fichier Excel (.xlsx/.xlsm/.xls) à charger automatiquement.
        Ignore les fichiers verrous (~$...) et cachés. Retourne le
        premier trouvé (ordre alphabétique), ou None si absent.
        """

        if not self.correlations_dir or not os.path.isdir(
            self.correlations_dir
        ):
            return None

        try:
            fichiers = sorted(os.listdir(self.correlations_dir))
        except Exception:
            return None

        for f in fichiers:
            if f.startswith("~$") or f.startswith("."):
                continue
            if f.lower().endswith((".xlsx", ".xlsm", ".xls")):
                return os.path.join(self.correlations_dir, f)

        return None

    # ======================================================

    def resizeEvent(self, event):

        self.bg.setGeometry(self.rect())

        if os.path.exists(BG):

            p = QPixmap(BG)

            if not p.isNull():

                self.bg.setPixmap(
                    p.scaled(
                        self.size(),
                        Qt.KeepAspectRatioByExpanding,
                        Qt.SmoothTransformation
                    )
                )

        super().resizeEvent(event)

    # ======================================================

    def build_ui(self):

        self.bg = QLabel(self)

        self.bg.setGeometry(self.rect())

        self.bg.setAlignment(Qt.AlignCenter)

        if os.path.exists(BG):

            p = QPixmap(BG)

            if not p.isNull():

                self.bg.setPixmap(
                    p.scaled(
                        self.size(),
                        Qt.KeepAspectRatioByExpanding,
                        Qt.SmoothTransformation
                    )
                )

        self.bg.lower()

        main = QVBoxLayout(self)

        main.setContentsMargins(20, 30, 20, 20)

        main.setSpacing(10)

        # Le grand titre décoratif "Albat / SCENAR" a été retiré :
        # l'onglet actif dans la barre d'onglets du haut suffit
        # déjà à indiquer où l'on se trouve.

        btn = QPushButton(
            "Choisir fichier Excel"
        )

        btn.clicked.connect(
            self.load_excel
        )

        main.addWidget(btn)

        self.file_label = QLabel(
            "Aucun fichier chargé"
        )

        self.file_label.setAlignment(Qt.AlignCenter)

        self.file_label.setStyleSheet("background:transparent; color:#4d4d4d;")

        main.addWidget(self.file_label)

        # ==========================================
        # VILLE DE RÉFÉRENCE (lever / coucher du soleil)
        # ==========================================

        lab_city = QLabel("Ville de référence (lever / coucher du soleil) :")
        lab_city.setStyleSheet("font-weight:bold; background:transparent;")
        main.addWidget(lab_city)

        city_line = QHBoxLayout()

        self.city_input = QLineEdit()
        self.city_input.setPlaceholderText("Ville de référence (ex : Dijon)")
        self.city_input.returnPressed.connect(self.search_city)

        btn_city = QPushButton("Rechercher")
        btn_city.setStyleSheet("""
            QPushButton{
                background:rgba(120,150,90,180);
                border:1px solid rgba(255,255,255,35);
                border-radius:12px;
                color:white;
                font-size:13px;
                font-weight:700;
                padding:8px 16px;
            }
            QPushButton:hover{
                background:rgba(160,210,120,220);
            }
            QPushButton:pressed{
                background:rgba(90,120,70,255);
            }
        """)
        btn_city.clicked.connect(self.search_city)

        city_line.addWidget(self.city_input)
        city_line.addWidget(btn_city)

        main.addLayout(city_line)

        self.city_status = QLabel("Aucune ville recherchée.")
        self.city_status.setWordWrap(True)
        self.city_status.setAlignment(Qt.AlignCenter)
        self.city_status.setStyleSheet("font-size:11px; background:transparent; color:#4d4d4d;")
        main.addWidget(self.city_status)

        if self.ref_city:
            self.city_input.setText(self.ref_city)

        if self.location_info is not None:
            self.city_status.setText(
                f"Trouvée : {self.location_info.name} "
                f"(lat {self.location_info.latitude:.2f}, "
                f"lon {self.location_info.longitude:.2f})"
            )

        self.scenario_name = QLineEdit()

        self.scenario_name.setPlaceholderText(
            "Nom du scénario de bridage"
        )

        main.addWidget(self.scenario_name)

        # ==========================================
        # ENREGISTRER / CHARGER UN PLAN
        # ==========================================

        plan_io_line = QHBoxLayout()
        plan_io_line.setSpacing(10)

        plan_io_style = """
            QPushButton{
                background:rgba(70,100,55,190);
                border:1px solid rgba(255,255,255,35);
                border-radius:10px;
                color:white;
                font-size:11px;
                font-weight:700;
                padding:10px 6px;
            }
            QPushButton:hover{
                background:rgba(100,140,75,220);
            }
        """

        save_plan_btn = QPushButton("Enregistrer le plan")
        save_plan_btn.setMinimumHeight(34)
        save_plan_btn.setStyleSheet(plan_io_style)
        save_plan_btn.clicked.connect(self.save_plan_to_file)

        load_plan_btn = QPushButton("Charger un plan")
        load_plan_btn.setMinimumHeight(34)
        load_plan_btn.setStyleSheet(plan_io_style)
        load_plan_btn.clicked.connect(self.load_plan_from_file)

        plan_io_line.addWidget(save_plan_btn, 1)
        plan_io_line.addWidget(load_plan_btn, 1)

        main.addLayout(plan_io_line)

        # ==========================================
        # TAUX DE PROTECTION GLOBAL DU PLAN
        # (combine toutes les plages, pas seulement une)
        # ==========================================

        self.lab_plan_pct = QLabel(
            "Taux de protection global du plan : —"
        )

        self.lab_plan_pct.setAlignment(Qt.AlignCenter)

        self.lab_plan_pct.setStyleSheet("""
            background: rgba(20,35,15,220);
            color: #f2e6d0;
            font-size: 16px;
            font-weight: 900;
            border-radius: 12px;
            padding: 10px;
            border: 2px solid rgba(160,210,120,160);
        """)

        main.addWidget(self.lab_plan_pct)

        # ==========================================
        # ONGLETS
        # ==========================================

        self.tabs = QTabWidget()

        self.tabs.setTabsClosable(True)

        self.tabs.tabCloseRequested.connect(
            self.close_tab
        )

        
        self.tabs.tabBarDoubleClicked.connect(
            self.rename_tab
        )

        self.tabs.currentChanged.connect(
            self.handle_tab
        )

        main.addWidget(self.tabs)

        # On bloque les signaux pendant toute l'initialisation pour
        # éviter qu'une bascule interne de l'onglet "+" ne déclenche
        # handle_tab -> add_tab en cascade (ce qui créait "Plage 2"
        # en trop dès l'ouverture).
        self.tabs.blockSignals(True)

        self.tabs.addTab(QWidget(), "+")
        self.add_tab()

        self.tabs.blockSignals(False)

        # Filet de sécurité : si plus d'une plage a été créée
        # pendant l'initialisation, on ne garde que la première.
        while self.tabs.count() > 2:
            extra_widget = self.tabs.widget(1)
            self.tabs.removeTab(1)
            if extra_widget is not None:
                extra_widget.deleteLater()

        # Sélectionner explicitement Plage 1
        self.tabs.setCurrentIndex(0)

        # ==========================================
        # EXPORT
        # ==========================================

        export_btn = QPushButton(
            "GENERER FICHIER EXCEL"
        )

        export_btn.clicked.connect(
            self.export_excel
        )

        main.addWidget(export_btn)

    # ======================================================

    def _collect_plan_data(self):
        """Construit le dictionnaire de données du plan (réutilisé
        par la sauvegarde manuelle et la sauvegarde de projet)."""

        data = {
            "nom": self.scenario_name.text().strip() or "Scenario Bridage",
            "ville_reference": self.city_input.text().strip(),
            "periodes": []
        }

        for i in range(self.tabs.count() - 1):

            tab = self.tabs.widget(i)

            data["periodes"].append({
                "nom": self.tabs.tabText(i),
                "debut": tab.date_start.date().toString("dd/MM/yyyy"),
                "fin": tab.date_end.date().toString("dd/MM/yyyy"),
                "vent_actif": tab.check_vent.isChecked(),
                "vent_max": tab.slider_vent.value() / 10,
                "temp_actif": tab.check_temp.isChecked(),
                "temp_min": tab.slider_temp.value() / 10,
                "nocturne_actif": tab.check_sun.isChecked(),
                "reference_debut": (
                    "coucher" if tab.start_ref.currentIndex() == 0
                    else "lever"
                ),
                "decalage_debut": tab.slider_start_offset.value(),
                "reference_fin": (
                    "lever" if tab.end_ref.currentIndex() == 0
                    else "coucher"
                ),
                "decalage_fin": tab.slider_end_offset.value(),
            })

        return data

    # ======================================================

    def save_plan_to_path(self, out_path):
        """Enregistre le plan directement vers un chemin donné,
        sans boîte de dialogue (utilisé par la sauvegarde globale
        du projet)."""

        import json

        data = self._collect_plan_data()

        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    # ======================================================

    def save_plan_to_file(self):
        """
        Enregistre le plan complet (nom du scénario, ville, et
        toutes les plages avec leurs réglages vent/température/
        plage nocturne) dans un fichier .json réutilisable.

        Utilise le même schéma que "Enregistrer les plans" dans
        Albat Bridage (clé "periodes", vent_max/temp_min en valeur
        réelle), pour permettre l'échange de fichiers entre les
        deux modules.
        """

        import json

        try:

            default_plan_name = "plan_scenar.json"

            if self.default_save_dir and os.path.isdir(self.default_save_dir):
                default_plan_name = os.path.join(
                    self.default_save_dir, default_plan_name
                )

            out_path, _ = QFileDialog.getSaveFileName(
                self,
                "Enregistrer le plan de protection",
                default_plan_name,
                "Fichier de plan (*.json)"
            )

            if not out_path:
                return

            if not out_path.lower().endswith(".json"):
                out_path += ".json"

            self.save_plan_to_path(out_path)

            data = self._collect_plan_data()

            QMessageBox.information(
                self, "Succès",
                f"Plan enregistré ({len(data['periodes'])} plage(s)) "
                f"dans :\n{out_path}"
            )

        except Exception as e:
            QMessageBox.critical(
                self, "Erreur",
                f"Impossible d'enregistrer le plan :\n{e}"
            )

    # ======================================================

    def _apply_plan_data(self, data):
        """Applique un dictionnaire de données de plan à
        l'interface (réutilisé par le chargement manuel et le
        chargement global du projet). Retourne (nb_plages, note)."""

        # Détection du format :
        # - fichier Scenar natif (clé "periodes" ou ancienne
        #   clé "plages" à la racine)
        # - fichier multi-plans Bridage (clé "plans") : on ne
        #   charge alors que le premier plan, Scenar ne gérant
        #   qu'un seul scénario à la fois.
        source_note = ""

        if "plans" in data:
            plans_list = data.get("plans", [])
            if not plans_list:
                raise Exception("Ce fichier ne contient aucun plan.")
            chosen = plans_list[0]
            nom = chosen.get("nom", "")
            periodes = chosen.get("periodes", [])
            if len(plans_list) > 1:
                source_note = (
                    f"\n(Ce fichier contient {len(plans_list)} plans "
                    f"Bridage ; seul le premier a été chargé.)"
                )
        else:
            nom = data.get("nom") or data.get("nom_scenario", "")
            periodes = data.get("periodes", data.get("plages", []))

        # Supprime toutes les plages actuelles (garde le "+")
        self.tabs.blockSignals(True)

        while self.tabs.count() > 1:
            widget = self.tabs.widget(0)
            self.tabs.removeTab(0)
            if widget is not None:
                widget.deleteLater()

        self.tabs.blockSignals(False)

        if nom:
            self.scenario_name.setText(nom)

        ville = data.get("ville_reference", "")
        if ville:
            self.city_input.setText(ville)

        for plage_data in periodes:

            self.add_tab()

            idx = self.tabs.count() - 2
            tab = self.tabs.widget(idx)

            self.tabs.setTabText(
                idx,
                plage_data.get("nom", self.tabs.tabText(idx))
            )

            debut = QDate.fromString(
                plage_data.get("debut", "01/04/2025"), "dd/MM/yyyy"
            )
            fin = QDate.fromString(
                plage_data.get("fin", "31/10/2025"), "dd/MM/yyyy"
            )

            if debut.isValid():
                tab.date_start.setDate(debut)
            if fin.isValid():
                tab.date_end.setDate(fin)

            tab.check_vent.setChecked(
                plage_data.get("vent_actif", True)
            )
            if "vent_max" in plage_data:
                tab.slider_vent.setValue(
                    round(plage_data.get("vent_max", 5.0) * 10)
                )
            else:
                tab.slider_vent.setValue(
                    plage_data.get("vent_valeur", 50)
                )

            tab.check_temp.setChecked(
                plage_data.get("temp_actif", True)
            )
            if "temp_min" in plage_data:
                tab.slider_temp.setValue(
                    round(plage_data.get("temp_min", 10.0) * 10)
                )
            else:
                tab.slider_temp.setValue(
                    plage_data.get("temp_valeur", 100)
                )

            tab.check_sun.setChecked(
                plage_data.get("nocturne_actif", False)
            )

            if "reference_debut" in plage_data:
                tab.start_ref.setCurrentIndex(
                    0 if plage_data.get("reference_debut") == "coucher"
                    else 1
                )
                tab.slider_start_offset.setValue(
                    plage_data.get("decalage_debut", 0)
                )
            else:
                # Compatibilité avec l'ancien schéma
                # (toujours coucher -> lever)
                tab.start_ref.setCurrentIndex(0)
                tab.slider_start_offset.setValue(
                    plage_data.get("decalage_coucher", 0)
                )

            if "reference_fin" in plage_data:
                tab.end_ref.setCurrentIndex(
                    0 if plage_data.get("reference_fin") == "lever"
                    else 1
                )
                tab.slider_end_offset.setValue(
                    plage_data.get("decalage_fin", 0)
                )
            else:
                tab.end_ref.setCurrentIndex(0)
                tab.slider_end_offset.setValue(
                    plage_data.get("decalage_lever", 0)
                )

        self.update_all_tabs()

        # Après reconstruction des onglets, chacun des add_tab()
        # ci-dessus a basculé l'onglet actif sur la plage qu'il
        # venait de créer : sans ce correctif, c'est donc la
        # dernière plage chargée qui reste sélectionnée à
        # l'ouverture du projet, pas "Plage 1".
        if self.tabs.count() > 0:
            self.tabs.setCurrentIndex(0)

        return len(periodes), source_note

    # ======================================================

    def load_plan_from_path(self, in_path):
        """Charge un plan directement depuis un chemin donné, sans
        boîte de dialogue (utilisé par le chargement global du
        projet)."""

        import json

        with open(in_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        return self._apply_plan_data(data)

    # ======================================================

    def load_plan_from_file(self):
        """
        Charge un plan précédemment enregistré, en remplaçant
        toutes les plages actuellement affichées.
        """

        try:

            in_path, _ = QFileDialog.getOpenFileName(
                self,
                "Charger un plan de protection",
                "",
                "Fichier de plan (*.json)"
            )

            if not in_path:
                return

            n_periodes, source_note = self.load_plan_from_path(in_path)

            QMessageBox.information(
                self, "Succès",
                f"Plan chargé ({n_periodes} plage(s)) "
                f"depuis :\n{in_path}{source_note}"
            )

        except Exception as e:
            QMessageBox.critical(
                self, "Erreur",
                f"Impossible de charger le plan :\n{e}"
            )

    # ======================================================

    def add_tab(self):

        page = PeriodTab(self)

        plus_index = self.tabs.count() - 1

        existing_names = []

        for i in range(self.tabs.count()-1):

            existing_names.append(
                self.tabs.tabText(i)
            )

        next_id = 1

        while f"Plage {next_id}" in existing_names:
            next_id += 1

        self.tabs.insertTab(
            plus_index,
            page,
            f"Plage {next_id}"
        )

        # Ne pas changer automatiquement d'onglet
        if self.tabs.count() > 2:
            self.tabs.setCurrentIndex(plus_index)

        self.refresh_names()

    # ======================================================

    def refresh_names(self):

        # Conserver les noms personnalisés existants
        # uniquement garantir le dernier onglet "+"
        if self.tabs.count() > 0:

            self.tabs.setTabText(
                self.tabs.count() - 1,
                "+"
            )

    # ======================================================


    # ======================================================

    def rename_tab(self, index):

        if index == self.tabs.count() - 1:
            return

        current_name = self.tabs.tabText(index)

        new_name, ok = QInputDialog.getText(
            self,
            "Renommer onglet",
            "Nom de la plage :",
            text=current_name
        )

        if ok and new_name.strip():

            self.tabs.setTabText(
                index,
                new_name.strip()
            )


    def handle_tab(self, index):

        # Ignore invalid transitions during close
        if index < 0:
            return

        # Onglet "+"
        if index == self.tabs.count()-1:

            self.add_tab()

    # ======================================================

    def close_tab(self, index):

        # Ne jamais fermer l'onglet "+"
        plus_index = self.tabs.count() - 1

        if index == plus_index:
            return

        # Garder au moins une plage
        if plus_index <= 1:
            return

        # Récupérer EXACTEMENT le widget ciblé
        widget_to_remove = self.tabs.widget(index)

        # Bloquer signaux pendant suppression
        self.tabs.blockSignals(True)

        # Retirer exactement cet onglet
        self.tabs.removeTab(index)

        self.tabs.blockSignals(False)

        # Détruire widget supprimé
        if widget_to_remove is not None:
            widget_to_remove.deleteLater()

        # Renommer uniquement les onglets restants
        self.refresh_names()

        # Sélection stable
        remaining_tabs = self.tabs.count() - 1

        if remaining_tabs > 0:

            new_index = min(index, remaining_tabs - 1)

            self.tabs.setCurrentIndex(new_index)

        self.update_all_tabs()

    # ======================================================

    def _lookup_city_geoapi(self, name):
        """
        Recherche une commune française via l'API officielle
        geo.api.gouv.fr (même méthode que Bridage / Graph).
        Retourne (nom, latitude, longitude) ou None.
        """

        import urllib.request
        import urllib.parse
        import json as _json

        try:
            query = urllib.parse.urlencode({
                "nom": name,
                "fields": "nom,centre,population",
                "boost": "population",
                "limit": 1,
            })
            url = f"https://geo.api.gouv.fr/communes?{query}"

            req = urllib.request.Request(
                url, headers={"User-Agent": "AlbatScenar/1.0"}
            )

            with urllib.request.urlopen(req, timeout=6) as response:
                data = _json.loads(response.read().decode("utf-8"))

            if not data:
                return None

            commune = data[0]
            lon, lat = commune["centre"]["coordinates"]

            return (commune["nom"], lat, lon)

        except Exception:
            return None

    # ======================================================

    def _resolve_ref_city(self):
        """
        Résout silencieusement la ville de référence transmise par
        la fenêtre principale (popup de démarrage), sans champ de
        recherche visible dans cet onglet.
        """

        result = self._lookup_city_geoapi(self.ref_city)

        if result is not None:

            found_name, lat, lon = result

            self.location_info = LocationInfo(
                name=found_name,
                region="France",
                timezone="Europe/Paris",
                latitude=lat,
                longitude=lon,
            )

    # ======================================================

    def get_sun_window(self, date):
        """
        Retourne (coucher, lever_lendemain) en UTC+2 fixe (naive),
        cohérent avec la convention utilisée dans Bridage.
        """

        if date in self._sun_cache:
            return self._sun_cache[date]

        tz = ZoneInfo(self.location_info.timezone)
        UTC = ZoneInfo("UTC")

        s_today = astral_sun(
            self.location_info.observer,
            date=date,
            tzinfo=tz
        )

        s_tomorrow = astral_sun(
            self.location_info.observer,
            date=date + dtmod.timedelta(days=1),
            tzinfo=tz
        )

        def to_utc2(aware_dt):
            return (
                aware_dt.astimezone(UTC).replace(tzinfo=None)
                + dtmod.timedelta(hours=2)
            )

        result = (
            to_utc2(s_today["sunset"]),
            to_utc2(s_tomorrow["sunrise"])
        )

        self._sun_cache[date] = result

        return result

    # ======================================================

    def compute_night_mask(
        self, df, date_col,
        start_ref, offset_start,
        end_ref, offset_end
    ):
        """
        Retourne un masque booléen : True si le contact tombe dans
        la plage définie par [référence début + décalage début,
        référence fin + décalage fin], où chaque référence est
        indépendamment "coucher" ou "lever" du soleil.

        Ex : start_ref="coucher", offset_start=-60 -> 1h avant le coucher.
             end_ref="coucher", offset_end=+180 -> 3h après le coucher.
        """

        if (
            not ASTRAL_OK
            or self.location_info is None
            or date_col is None
        ):
            return pd.Series([False] * len(df), index=df.index)

        results = []

        for dt in df[date_col]:

            if pd.isna(dt):
                results.append(False)
                continue

            dt = dt.to_pydatetime() if hasattr(dt, "to_pydatetime") else dt

            ref_date = dt.date()

            if dt.time() < dtmod.time(12, 0):
                ref_date = ref_date - dtmod.timedelta(days=1)

            try:
                sunset_t, sunrise_t = self.get_sun_window(ref_date)
            except Exception:
                results.append(False)
                continue

            start_base = sunset_t if start_ref == "coucher" else sunrise_t
            end_base = sunset_t if end_ref == "coucher" else sunrise_t

            window_start = start_base + dtmod.timedelta(minutes=offset_start)
            window_end = end_base + dtmod.timedelta(minutes=offset_end)

            results.append(window_start <= dt <= window_end)

        return pd.Series(results, index=df.index)

    # ======================================================

    def search_city(self):

        name = self.city_input.text().strip()

        if not name:
            self.city_status.setText("Veuillez saisir un nom de ville.")
            return

        if not ASTRAL_OK:
            self.city_status.setText(
                "Module 'astral' manquant : exécutez 'pip install astral'."
            )
            return

        self.city_status.setText("Recherche en ligne...")
        QApplication.processEvents()

        result = self._lookup_city_geoapi(name)

        # La ville change : les heures de lever/coucher déjà mises
        # en cache (pour l'ancienne ville) ne sont plus valables.
        self._sun_cache = {}

        if result is not None:

            found_name, lat, lon = result

            self.location_info = LocationInfo(
                name=found_name,
                region="France",
                timezone="Europe/Paris",
                latitude=lat,
                longitude=lon,
            )

            self.city_status.setText(
                f"Trouvée : {found_name} "
                f"(lat {lat:.2f}, lon {lon:.2f})"
            )

        else:

            self.location_info = None

            self.city_status.setText(
                "Ville introuvable. Vérifiez l'orthographe, ou votre "
                "connexion internet."
            )

        # La ville a changé : on rafraîchit immédiatement les taux
        # de protection si un fichier est déjà chargé.
        if not self.df.empty:
            self.update_all_tabs()

    # ======================================================

    def load_excel(self):

        file, _ = QFileDialog.getOpenFileName(
            self,
            "Excel",
            "",
            "Excel (*.xlsx *.xls *.xlsm)"
        )

        if not file:
            return

        self.df = pd.read_excel(file)

        self.file_label.setText(
            os.path.basename(file)
        )

        self.update_all_tabs()

    # ======================================================

    def detect_species_col(self):

        for c in self.df.columns:

            lc = str(c).lower()

            if (
                "espèce" in lc
                or "espece" in lc
                or "species" in lc
            ):
                return c

        return None

    # ======================================================

    def detect_vent_col(self):

        for c in self.df.columns:

            lc = str(c).strip().lower()

            if lc in ("vent", "wind", "vitesse du vent", "ws"):
                return c

        return None

    # ======================================================

    def detect_temp_col(self):

        for c in self.df.columns:

            lc = str(c).strip().lower()

            if lc in ("température", "temperature", "temp"):
                return c

        return None

    # ======================================================

    def update_all_tabs(self):

        if self.df.empty:
            return

        species_col = self.detect_species_col()

        if species_col is None:
            return

        vent_col = self.detect_vent_col()
        temp_col = self.detect_temp_col()

        if vent_col is None or temp_col is None:
            QMessageBox.warning(
                self, "Colonnes introuvables",
                "Colonne vent et/ou température introuvable dans le "
                "fichier chargé (attendu : 'vent' et 'température')."
            )
            return

        plan_protected_indices = set()

        for i in range(self.tabs.count()-1):

            tab = self.tabs.widget(i)

            df = self.df.copy()

            df = df.rename(columns={
                vent_col: "vent",
                temp_col: "temp",
            })

            date_col = None

            for c in df.columns:
                if str(c).strip().lower() == "date_heure":
                    date_col = c
                    break

            if date_col is None:
                for c in df.columns:
                    if "date" in str(c).lower():
                        date_col = c
                        break

            if date_col is not None:

                df[date_col] = pd.to_datetime(
                    df[date_col],
                    errors="coerce",
                    dayfirst=True
                )

                start = tab.date_start.date().toPython()
                end = tab.date_end.date().toPython()

                df = df[
                    (
                        df[date_col].dt.date >= start
                    )
                    &
                    (
                        df[date_col].dt.date <= end
                    )
                ]

            protected = df[
                tab.compute_mask(df, date_col)
            ]

            plan_protected_indices.update(protected.index)

            species = sorted(
                df[species_col]
                .dropna()
                .unique()
            )

            # ======================================
            # +1 ligne TOTAL GLOBAL
            # ======================================

            tab.table.setRowCount(len(species) + 1)

            total_all = len(df)

            protected_all = len(protected)

            pct_all = 0

            if total_all > 0:
                pct_all = (
                    protected_all
                    / total_all
                ) * 100

            tab.lab_global_pct.setText(
                f"Taux de protection de la plage : {round(pct_all, 1)}% "
                f"({protected_all}/{total_all} contacts)"
            )

            total_vals = [
                "TOTAL GLOBAL",
                str(protected_all),
                str(total_all),
                f"{round(pct_all,1)}%"
            ]

            for c, v in enumerate(total_vals):

                item = QTableWidgetItem(v)

                item.setTextAlignment(Qt.AlignCenter)

                font = item.font()
                font.setBold(True)

                item.setFont(font)

                tab.table.setItem(0, c, item)

            for r, sp in enumerate(species, start=1):

                total = len(
                    df[
                        df[species_col] == sp
                    ]
                )

                prot = len(
                    protected[
                        protected[species_col] == sp
                    ]
                )

                pct = 0

                if total > 0:
                    pct = (prot / total) * 100

                vals = [
                    str(sp),
                    str(prot),
                    str(total),
                    f"{round(pct,1)}%"
                ]

                for c, v in enumerate(vals):

                    item = QTableWidgetItem(v)

                    item.setTextAlignment(Qt.AlignCenter)

                    if c == 0:
                        item.setToolTip(v)

                    tab.table.setItem(r, c, item)

        # ======================================
        # TAUX DE PROTECTION GLOBAL DU PLAN
        # (toutes les plages combinées, sans double
        # comptage si des plages se chevauchent)
        # ======================================

        total_plan = len(self.df)
        protected_plan = len(plan_protected_indices)

        pct_plan = 0

        if total_plan > 0:
            pct_plan = (protected_plan / total_plan) * 100

        self.lab_plan_pct.setText(
            f"Taux de protection global du plan : "
            f"{round(pct_plan, 1)}% "
            f"({protected_plan}/{total_plan} contacts)"
        )


    # ======================================================

    def _style_worksheet(self, ws):
        """
        Applique un style homogène à une feuille de données :
        en-tête vert avec texte blanc, colonnes dimensionnées
        selon leur contenu, formats de date/heure corrects,
        police cohérente, et ligne d'en-tête figée.
        Même style que Correlations et Bridage.
        """

        header_font = Font(name="Arial", bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4C6B3A")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )

        for col_idx, col_cells in enumerate(
            ws.iter_cols(min_row=1, max_row=ws.max_row), start=1
        ):
            header_text = str(col_cells[0].value or "")
            max_len = len(header_text)
            sample_val = None

            for cell in col_cells[1:]:

                if cell.value is not None:

                    if sample_val is None:
                        sample_val = cell.value

                    length = len(str(cell.value))

                    if length > max_len:
                        max_len = length

                if cell.font is None or cell.font.name != "Arial":
                    cell.font = Font(name="Arial")

            width = min(max(max_len + 3, 10), 34)
            ws.column_dimensions[
                get_column_letter(col_idx)
            ].width = width

            if isinstance(sample_val, dtmod.datetime):

                is_full_datetime = any(
                    c.value is not None
                    and isinstance(c.value, dtmod.datetime)
                    and c.value.time() != dtmod.time(0, 0)
                    for c in col_cells[1:200]
                )

                fmt = (
                    "dd/mm/yyyy hh:mm:ss"
                    if is_full_datetime
                    else "dd/mm/yyyy"
                )

                for cell in col_cells[1:]:
                    if cell.value is not None:
                        cell.number_format = fmt

            elif isinstance(sample_val, dtmod.date):

                for cell in col_cells[1:]:
                    if cell.value is not None:
                        cell.number_format = "dd/mm/yyyy"

        ws.freeze_panes = "A2"

    # ======================================================

    def _build_synthese_sheet(self, wb, scenario_name, plage_infos, global_stats):
        """
        Ajoute un onglet 'Synthèse' récapitulant le scénario :
        ses plages (dates, seuils vent/température, restriction
        nocturne) et le taux de protection par plage ainsi que le
        taux de protection global du plan. Même mise en forme que
        l'onglet 'Synthèse des plans' de Bridage.
        """

        ws = wb.create_sheet("Synthèse")

        title_font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
        title_fill = PatternFill("solid", fgColor="4C6B3A")

        plan_font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        plan_fill = PatternFill("solid", fgColor="6E8F4E")

        table_header_font = Font(name="Arial", bold=True, color="FFFFFF")
        table_header_fill = PatternFill("solid", fgColor="8FAE6B")

        normal_font = Font(name="Arial", size=10)

        row = 1

        ws.cell(row=row, column=1, value="SYNTHÈSE DU PLAN DE BRIDAGE")
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=5
        )
        cell = ws.cell(row=row, column=1)
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

        info_year = self.ref_year if self.ref_year else dtmod.date.today().year
        info_park = self.park_name if self.park_name else "-"

        ws.cell(
            row=row, column=1,
            value=f"Parc éolien : {info_park}    |    Année : {info_year}"
        )
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=5
        )
        cell = ws.cell(row=row, column=1)
        cell.font = Font(name="Arial", italic=True, size=10, color="444444")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 2

        ws.cell(row=row, column=1, value=scenario_name)
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=5
        )
        cell = ws.cell(row=row, column=1)
        cell.font = plan_font
        cell.fill = plan_fill
        cell.alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        ws.row_dimensions[row].height = 20
        row += 1

        headers = [
            "Plage", "Période", "Vent <", "Temp >",
            "Restriction nocturne", "Taux de protection"
        ]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = table_header_font
            cell.fill = table_header_fill
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Largeur de chaque colonne calculée d'après le texte le
        # plus long qu'elle contiendra réellement (en-tête compris)
        # plutôt qu'une largeur fixe : une période complète type
        # "01/04/2025 → 30/09/2025" (24 caractères) ne rentrait pas
        # dans les 20 caractères fixes précédemment alloués à la
        # colonne "Période", ce qui la tronquait visuellement.
        col_max_lens = [len(h) for h in headers]

        for info in plage_infos:

            row_values = [
                str(info["nom"]),
                f"{info['debut']} → {info['fin']}",
                str(info["vent_txt"]),
                str(info["temp_txt"]),
                str(info["sun_txt"]),
                str(info["taux_txt"]),
            ]

            for c, val in enumerate(row_values):
                col_max_lens[c] = max(col_max_lens[c], len(val))

            ws.cell(row=row, column=1, value=info["nom"]).font = normal_font
            ws.cell(
                row=row, column=2,
                value=f"{info['debut']} → {info['fin']}"
            ).font = normal_font
            ws.cell(row=row, column=3, value=info["vent_txt"]).font = normal_font
            ws.cell(row=row, column=4, value=info["temp_txt"]).font = normal_font
            ws.cell(row=row, column=5, value=info["sun_txt"]).font = normal_font
            ws.cell(
                row=row, column=6, value=info["taux_txt"]
            ).font = Font(name="Arial", size=10, bold=True)

            for c in range(1, 7):
                ws.cell(row=row, column=c).alignment = Alignment(
                    horizontal="center"
                )

            row += 1

        row += 1

        ws.cell(
            row=row, column=1,
            value="TAUX DE PROTECTION GLOBAL DU PLAN"
        )
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=4
        )
        cell = ws.cell(row=row, column=1)
        cell.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4C6B3A")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.cell(row=row, column=5, value=global_stats["taux_txt"])
        ws.merge_cells(
            start_row=row, start_column=5, end_row=row, end_column=6
        )
        cell = ws.cell(row=row, column=5)
        cell.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4C6B3A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 2

        ws.cell(
            row=row, column=1,
            value=(
                f"Contacts protégés : {global_stats['protected']} / "
                f"{global_stats['total']}"
            )
        ).font = normal_font

        for col_letter, max_len in zip("ABCDEF", col_max_lens):
            # Marge de confort (+4) et plancher à 12 pour rester
            # lisible même sur une colonne à contenu très court.
            ws.column_dimensions[col_letter].width = max(
                max_len + 4, 12
            )

    # ======================================================

    def export_excel(self):

        if self.df.empty:
            return

        out_year = self.ref_year if self.ref_year else dtmod.date.today().year
        park_part = self.park_name if self.park_name else "Albat"

        default_name = f"{park_part}_{out_year}_Scenar.xlsx"

        if self.default_save_dir and os.path.isdir(self.default_save_dir):
            default_name = os.path.join(self.default_save_dir, default_name)

        out, _ = QFileDialog.getSaveFileName(
            self,
            "Excel",
            default_name,
            "Excel (*.xlsx)"
        )

        if not out:
            return

        try:

            with pd.ExcelWriter(out, engine="openpyxl") as writer:

                params = []
                plage_infos = []

                # ==================================================
                # EXPORT DES FEUILLES PAR PLAGE
                # ==================================================

                for i in range(self.tabs.count()-1):

                    tab = self.tabs.widget(i)

                    data = []

                    for row in range(tab.table.rowCount()):

                        vals = []

                        for col in range(tab.table.columnCount()):

                            item = tab.table.item(row, col)

                            if item:
                                vals.append(item.text())
                            else:
                                vals.append("")

                        data.append(vals)

                    df_export = pd.DataFrame(
                        data,
                        columns=[
                            "Espèce",
                            "Contacts protégés",
                            "Total",
                            "% Protection"
                        ]
                    )

                    sheet_name = self.tabs.tabText(i)[:31]

                    df_export.to_excel(
                        writer,
                        sheet_name=sheet_name,
                        index=False
                    )

                    self._style_worksheet(writer.sheets[sheet_name])

                    params.append({
                        "Nom plage": self.tabs.tabText(i),
                        "Date début": tab.date_start.date().toString("dd/MM/yyyy"),
                        "Date fin": tab.date_end.date().toString("dd/MM/yyyy"),
                        "Vent actif": "Oui" if tab.check_vent.isChecked() else "Non",
                        "Vent max": f"{tab.slider_vent.value()/10:.1f} m/s",
                        "Temp actif": "Oui" if tab.check_temp.isChecked() else "Non",
                        "Temp min": f"{tab.slider_temp.value()/10:.1f} °C"
                    })

                    # Infos pour l'onglet Synthèse (lues depuis la ligne
                    # "TOTAL GLOBAL" déjà calculée par le tableau à l'écran)
                    prot_txt = data[0][1] if data else "0"
                    total_txt = data[0][2] if data else "0"
                    pct_txt = data[0][3] if data else "0%"

                    plage_infos.append({
                        "nom": self.tabs.tabText(i),
                        "debut": tab.date_start.date().toString("dd/MM/yyyy"),
                        "fin": tab.date_end.date().toString("dd/MM/yyyy"),
                        "vent_txt": (
                            f"< {tab.slider_vent.value()/10:.1f} m/s"
                            if tab.check_vent.isChecked() else "Sans limite"
                        ),
                        "temp_txt": (
                            f"> {tab.slider_temp.value()/10:.1f} °C"
                            if tab.check_temp.isChecked() else "Sans limite"
                        ),
                        "sun_txt": (
                            f"Début: {tab.start_ref.currentText()} "
                            f"{tab.slider_start_offset.value():+d}min → "
                            f"Fin: {tab.end_ref.currentText()} "
                            f"{tab.slider_end_offset.value():+d}min"
                            if tab.check_sun.isChecked() else "Aucune (jour et nuit)"
                        ),
                        "taux_txt": f"{pct_txt} ({prot_txt}/{total_txt})",
                    })

                # ==================================================
                # DONNEES COMPLETES
                # ==================================================

                full_export = self.df.copy()

                vent_col = self.detect_vent_col()
                temp_col = self.detect_temp_col()

                global_mask = pd.Series(
                    [False] * len(full_export),
                    index=full_export.index
                )

                for i in range(self.tabs.count()-1):

                    tab = self.tabs.widget(i)

                    df_period = self.df.copy()

                    if vent_col is not None and temp_col is not None:
                        df_period = df_period.rename(columns={
                            vent_col: "vent",
                            temp_col: "temp",
                        })

                    date_col = None

                    for c in df_period.columns:
                        if str(c).strip().lower() == "date_heure":
                            date_col = c
                            break

                    if date_col is None:
                        for c in df_period.columns:
                            if "date" in str(c).lower():
                                date_col = c
                                break

                    if date_col is not None:

                        df_period[date_col] = pd.to_datetime(
                            df_period[date_col],
                            errors="coerce",
                            dayfirst=True
                        )

                        start = tab.date_start.date().toPython()
                        end = tab.date_end.date().toPython()

                        date_mask = (
                            (
                                df_period[date_col].dt.date >= start
                            )
                            &
                            (
                                df_period[date_col].dt.date <= end
                            )
                        )

                    else:

                        date_mask = pd.Series(
                            [True] * len(df_period),
                            index=df_period.index
                        )

                    final_mask = (
                        date_mask
                        &
                        tab.compute_mask(df_period, date_col)
                    )

                    global_mask = (
                        global_mask
                        |
                        final_mask
                    )

                scenario_name = self.scenario_name.text().strip()

                if scenario_name == "":
                    scenario_name = "Scenario Bridage"

                full_export[scenario_name] = global_mask.map(
                    lambda x: "Oui" if x else "Non"
                )

                full_export.to_excel(
                    writer,
                    sheet_name="Données Complètes",
                    index=False
                )

                self._style_worksheet(writer.sheets["Données Complètes"])

                # ==================================================
                # FEUILLE PARAMETRES
                # ==================================================

                params_df = pd.DataFrame(params)

                params_df.to_excel(
                    writer,
                    sheet_name="Paramètres",
                    index=False
                )

                self._style_worksheet(writer.sheets["Paramètres"])

                # ==================================================
                # ONGLET SYNTHÈSE
                # ==================================================

                total_global = len(full_export)
                protected_global = int(global_mask.sum())

                pct_global = (
                    round(protected_global / total_global * 100, 1)
                    if total_global > 0 else 0
                )

                global_stats = {
                    "protected": protected_global,
                    "total": total_global,
                    "taux_txt": f"{pct_global}% ({protected_global}/{total_global})",
                }

                self._build_synthese_sheet(
                    writer.book, scenario_name, plage_infos, global_stats
                )

                wb = writer.book
                wb.move_sheet("Synthèse", offset=-(len(wb.sheetnames) - 1))

                # ==================================================
                # COLORATION OPENPYXL
                # ==================================================

                ws = writer.sheets["Données Complètes"]

                green_fill = PatternFill(
                    fill_type="solid",
                    start_color="92D050",
                    end_color="92D050"
                )

                red_fill = PatternFill(
                    fill_type="solid",
                    start_color="FF7C80",
                    end_color="FF7C80"
                )

                scenario_col = len(full_export.columns)

                for row in range(2, len(full_export) + 2):

                    value = ws.cell(
                        row=row,
                        column=scenario_col
                    ).value

                    fill = (
                        green_fill
                        if str(value).strip().lower() == "oui"
                        else red_fill
                    )

                    for col in range(1, scenario_col + 1):

                        ws.cell(
                            row=row,
                            column=col
                        ).fill = fill

        except Exception as e:
            QMessageBox.critical(
                self, "Erreur",
                f"Impossible de générer le fichier Excel :\n{e}"
            )
            return
        QMessageBox.information(
            self,
            "Succès",
            "Fichier Excel généré."
        )



if __name__ == "__main__":

    app = QApplication(sys.argv)

    win = OptimisationWindow()

    win.show()

    sys.exit(app.exec())
