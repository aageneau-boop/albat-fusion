# ==========================================================
# BatCorrelations - VERSION PROPRE pandas récent
# Compatible :
# - pandas récent
# - Vilpion
# - SCADA
# - Montie
# - Corrélations réelles
# ==========================================================

import sys
import os
import pandas as pd
import openpyxl
import datetime as dtmod
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import subprocess
def open_file(path):
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.call(["open", path])
        else:
            subprocess.call(["xdg-open", path])
    except Exception:
        pass


from PySide6.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton,
    QFileDialog, QVBoxLayout, QHBoxLayout, QProgressBar,
    QMessageBox, QComboBox, QFrame
)

from PySide6.QtGui import QPixmap, QIcon
from PySide6.QtCore import Qt, QSize


# ----------------------------------------------------------
# chemins
# ----------------------------------------------------------
from modules.utils import resource_path


ASSETS = resource_path("assets")

BG = os.path.join(ASSETS, "background_graph.png")
LOGO = os.path.join(ASSETS, "logo_bat.png")
ICON_EXCEL = os.path.join(ASSETS, "icon_excel.png")
ICON_WIND = os.path.join(ASSETS, "icon_wind.png")
ICON_RUN = os.path.join(ASSETS, "icon_launch.png")


# ----------------------------------------------------------
class BatCorrelations(QWidget):


    # ------------------------------------------------------
    def resizeEvent(self, event):

        if hasattr(self, "bg") and self.bg:

            self.bg.setGeometry(self.rect())

            if os.path.exists(BG):

                p = QPixmap(BG)

                if not p.isNull():

                    self.bg.setPixmap(
                        p.scaled(
                            self.size(),
                            Qt.KeepAspectRatioByExpanding,
                            Qt.SmoothTransformation
                        )
                    )

        super().resizeEvent(event)

    def __init__(self, park_name="", ref_year=None, bureau_etude="", default_save_dir=""):
        super().__init__()

        self.park_name = park_name
        self.ref_year = ref_year
        self.bureau_etude = bureau_etude
        self.default_save_dir = default_save_dir
        self.fichier_excel = ""
        self.fichier_csv = ""

        self.setWindowTitle("Albat Correlations")
        self.resize(500,900)
        icon_path = os.path.join(ASSETS, "icone_albat.ico")

        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        self.setStyleSheet("""
        QWidget{
            background:#071107;
            color:white;
            font-family:Cinzel SemiBold;
        }

        QPushButton{
    background:rgba(90,127,71,145);
    border:1px solid rgba(255,255,255,40);
    border-radius:12px;
    padding:6px 10px;
    text-align:center;
    font-size:13px;
    font-weight:700;
    color:white;
}

QPushButton:hover{
    background:rgba(130,184,99,220);
}

        QPushButton#run{
            font-size:24px;
            text-align:center;
            padding:16px;
        }

        QLabel#box{
            background:rgba(0,0,0,120);
            border-radius:14px;
            padding:12px;
        }

        QProgressBar{
            background:#081008;
            border-radius:11px;
            height:22px;
        }

        QProgressBar::chunk{
            background:#a6f56f;
            border-radius:11px;
        }
        """)

        self.ui()

    # ------------------------------------------------------
    def ui(self):

        self.bg = QLabel(self)
        self.bg.setGeometry(0, 0, 500, 900)
        self.bg.setAlignment(Qt.AlignCenter)
        if os.path.exists(BG):
            p = QPixmap(BG)
            if not p.isNull():
                self.bg.setPixmap(
                    p.scaled(
                        self.size(),
                        Qt.KeepAspectRatioByExpanding,
                        Qt.SmoothTransformation
                    )
                )

        self.bg.lower()
        

        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 20, 20, 20)
        lay.setSpacing(14)
        
        

        logo = QLabel()
        logo.setAlignment(Qt.AlignCenter)
        logo.setStyleSheet("background:transparent;")

        if os.path.exists(LOGO):
            p = QPixmap(LOGO)
            if not p.isNull():
                logo.setPixmap(
                    p.scaled(
                        140, 140,
                        Qt.KeepAspectRatio,
                        Qt.SmoothTransformation
                    )
                )

        # Le grand titre décoratif "Albat / CORRELATIONS" a été
        # retiré : l'onglet actif dans la barre d'onglets du haut
        # suffit déjà à indiquer où l'on se trouve.

        sous = QLabel("Corrélations entre présence de chiroptères et météo")
        sous.setAlignment(Qt.AlignCenter)
        sous.setStyleSheet("""
            background: transparent;
            color: #ffffff;
        """)

        self.lbl1 = QLabel("Aucun fichier brut sélectionné")
        self.lbl1.setObjectName("box")

        self.lbl2 = QLabel("Aucun fichier météo sélectionné")
        self.lbl2.setObjectName("box")

        b1 = QPushButton("Choisir fichier brut (.xlsx)")
        b1.clicked.connect(self.pick_excel)
        b1.setMinimumHeight(38)

        lbl_tz1 = QLabel("Fuseau horaire du fichier brut :")
        lbl_tz1.setStyleSheet("background:transparent;")

        self.tz_excel = QComboBox()
        self.tz_excel.addItems(["UTC+2", "UTC"])
        self.tz_excel.setFixedWidth(320)

        b2 = QPushButton("Choisir fichier météo (.csv)")
        b2.clicked.connect(self.pick_csv)
        b2.setMinimumHeight(38)

        lbl_tz2 = QLabel("Fuseau horaire du fichier météo :")
        lbl_tz2.setStyleSheet("background:transparent;")

        self.tz_csv = QComboBox()
        self.tz_csv.addItems(["UTC+2", "UTC"])
        self.tz_csv.setFixedWidth(320)

        run = QPushButton("CORRELER")
        run.setObjectName("run")
        run.clicked.connect(self.lancer_analyse)
        run.setMinimumHeight(52)
        run.setFixedWidth(260)

        self.pb = QProgressBar()
        self.pb.setValue(0)

        b1.setFixedWidth(320)
        b2.setFixedWidth(320)
        run.setFixedWidth(320)

        brut_card = QFrame()
        brut_card.setStyleSheet("""
            QFrame{
                background: rgba(90,127,71,60);
                border: 1px solid rgba(255,255,255,35);
                border-radius: 14px;
            }
        """)
        brut_layout = QVBoxLayout(brut_card)
        brut_layout.setSpacing(4)
        brut_layout.setContentsMargins(10, 8, 10, 8)

        brut_layout.addWidget(b1, alignment=Qt.AlignCenter)
        brut_layout.addWidget(self.lbl1)
        brut_layout.addWidget(lbl_tz1, alignment=Qt.AlignCenter)
        brut_layout.addWidget(self.tz_excel, alignment=Qt.AlignCenter)

        meteo_card = QFrame()
        meteo_card.setStyleSheet("""
            QFrame{
                background: rgba(90,127,71,60);
                border: 1px solid rgba(255,255,255,35);
                border-radius: 14px;
            }
        """)
        meteo_layout = QVBoxLayout(meteo_card)
        meteo_layout.setSpacing(4)
        meteo_layout.setContentsMargins(10, 8, 10, 8)

        meteo_layout.addWidget(b2, alignment=Qt.AlignCenter)
        meteo_layout.addWidget(self.lbl2)
        meteo_layout.addWidget(lbl_tz2, alignment=Qt.AlignCenter)
        meteo_layout.addWidget(self.tz_csv, alignment=Qt.AlignCenter)

        # Bloc "fichier brut" ancré tout en haut, un espace
        # élastique qui absorbe la place disponible, puis le bloc
        # "météo" qui se retrouve poussé vers le bas (comme avant),
        # suivi du sous-titre juste au-dessus du bouton CORRELER.
        lay.addWidget(brut_card)
        lay.addStretch()

        lay.addWidget(meteo_card)

        lay.addSpacing(20)

        lay.addWidget(sous, alignment=Qt.AlignCenter)
        lay.addSpacing(8)

        lay.addWidget(run, alignment=Qt.AlignCenter)

        lay.addWidget(self.pb)
    # ------------------------------------------------------
    def pick_excel(self):

        f, _ = QFileDialog.getOpenFileName(
            self,
            "Choisir fichier brut",
            "",
            "Excel (*.xlsx *.xlsm)"
        )

        if f:
            self.fichier_excel = f
            self.lbl1.setText(os.path.basename(f))

    # ------------------------------------------------------
    def pick_csv(self):

        f, _ = QFileDialog.getOpenFileName(
            self,
            "Choisir fichier météo",
            "",
            "CSV (*.csv)"
        )

        if f:
            self.fichier_csv = f
            self.lbl2.setText(os.path.basename(f))

    # ------------------------------------------------------

    def _style_worksheet(self, ws):
        """
        Applique un style homogène à une feuille de données :
        en-tête vert avec texte blanc, colonnes dimensionnées
        selon leur contenu, formats de date/heure corrects,
        police cohérente, et ligne d'en-tête figée.
        """

        header_font = Font(name="Arial", bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4C6B3A")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )

        for col_idx, col_cells in enumerate(
            ws.iter_cols(min_row=1, max_row=ws.max_row), start=1
        ):
            header_text = str(col_cells[0].value or "")
            max_len = len(header_text)
            sample_val = None

            for cell in col_cells[1:]:

                if cell.value is not None:

                    if sample_val is None:
                        sample_val = cell.value

                    length = len(str(cell.value))

                    if length > max_len:
                        max_len = length

                cell.font = Font(name="Arial")
                cell.alignment = Alignment(horizontal="center")

            width = min(max(max_len + 3, 10), 34)
            ws.column_dimensions[
                get_column_letter(col_idx)
            ].width = width

            if isinstance(sample_val, dtmod.datetime):

                is_full_datetime = any(
                    c.value is not None
                    and isinstance(c.value, dtmod.datetime)
                    and c.value.time() != dtmod.time(0, 0)
                    for c in col_cells[1:200]
                )

                fmt = (
                    "dd/mm/yyyy hh:mm:ss"
                    if is_full_datetime
                    else "dd/mm/yyyy"
                )

                for cell in col_cells[1:]:
                    if cell.value is not None:
                        cell.number_format = fmt

            elif isinstance(sample_val, dtmod.date):

                for cell in col_cells[1:]:
                    if cell.value is not None:
                        cell.number_format = "dd/mm/yyyy"

        ws.freeze_panes = "A2"

    # ------------------------------------------------------

    def _hide_redundant_columns(self, ws, header_names):
        """
        Masque (sans les supprimer) les colonnes dont l'en-tête
        correspond exactement à l'un des noms donnés. Les données
        restent présentes dans le fichier — utile ici pour cacher
        'Date' et 'Heure', redondantes avec 'date_heure', sans
        casser les modules (Graph) qui en ont besoin.
        """

        for col_idx, cell in enumerate(ws[1], start=1):

            header_text = str(cell.value or "").strip().lower()

            if header_text in [h.lower() for h in header_names]:
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].hidden = True

    # ------------------------------------------------------
    def lancer_analyse(self):

        try:
            self.pb.setValue(10)

            # ----------------------------------
            # EXCEL
            # ----------------------------------
            df = pd.read_excel(self.fichier_excel)

            col_date = None
            col_heure = None
            col_contacts = None

            for c in df.columns:

                lc = c.lower().strip()

                if lc == "date":
                    col_date = c

                elif lc == "heure":
                    col_heure = c

                elif "contact" in lc:
                    col_contacts = c

            if not col_date or not col_heure:
                raise Exception(
                    "Colonnes Date / Heure introuvables."
                )

            df["date_heure"] = pd.to_datetime(
                df[col_date].astype(str)
                + " "
                + df[col_heure].astype(str),
                errors="coerce",
                dayfirst=True
            )

            df = df.dropna(subset=["date_heure"])

            # Conversion en UTC+2 si le fichier brut est fourni en UTC
            if self.tz_excel.currentText() == "UTC":
                df["date_heure"] = df["date_heure"] + pd.Timedelta(hours=2)

            self.pb.setValue(30)

            # ----------------------------------
            # CSV météo auto
            # ----------------------------------
            met = pd.read_csv(
                self.fichier_csv,
                sep=None,
                engine="python",
                encoding="utf-8",
                on_bad_lines="skip"
            )

            met.columns = [str(c).strip() for c in met.columns]

            col_time = None
            col_ws = None
            col_temp = None

            for c in met.columns:

                lc = c.lower().strip()

                if (
                    "time" in lc
                    or "stamp" in lc
                    or "date" in lc
                ) and col_time is None:
                    col_time = c

                if (
                    "ws" in lc
                    or "wind" in lc
                    or "vent" in lc
                ) and col_ws is None:
                    col_ws = c

                if "temp" in lc and col_temp is None:
                    col_temp = c

            if not col_time:
                col_time = met.columns[0]

            if not col_ws:
                col_ws = met.columns[1]

            if not col_temp:
                col_temp = met.columns[2]

            met["date_heure"] = pd.to_datetime(
                met[col_time].astype(str).str.strip(),
                errors="coerce",
                format="mixed",
                dayfirst=True
            )

            met["vent"] = pd.to_numeric(
                met[col_ws],
                errors="coerce"
            )

            met["temp"] = pd.to_numeric(
                met[col_temp],
                errors="coerce"
            )

            met = met.dropna(subset=["date_heure"])

            if met.empty:
                raise Exception(
                    "Aucune date valide dans le fichier météo."
                )

            # Conversion en UTC+2 si le fichier météo est fourni en UTC
            if self.tz_csv.currentText() == "UTC":
                met["date_heure"] = met["date_heure"] + pd.Timedelta(hours=2)

            self.pb.setValue(55)

            # ----------------------------------
            # matching nearest global
            # ----------------------------------
            vents = []
            temps = []
            ecarts = []

            total = len(df)

            for i, dt in enumerate(df["date_heure"]):

                diff = (met["date_heure"] - dt).abs()

                if diff.empty:
                    vents.append(None)
                    temps.append(None)
                    ecarts.append(None)
                    continue

                idx = diff.idxmin()

                delta = diff.loc[idx]

                if delta <= pd.Timedelta(minutes=10):

                    vents.append(
                        met.loc[idx, "vent"]
                    )

                    temps.append(
                        met.loc[idx, "temp"]
                    )

                    ecarts.append(
                        round(delta.total_seconds() / 60, 1)
                    )

                else:

                    vents.append(None)
                    temps.append(None)
                    ecarts.append(None)

                self.pb.setValue(
                    55 + int((i + 1) / total * 25)
                )

            df["vent"] = vents
            df["température"] = temps
            df["Écart SCADA (min)"] = ecarts

                        # ----------------------------------
                        # corrélations
                        # ----------------------------------
            resume = pd.DataFrame()

            if col_contacts and col_contacts in df.columns:

                df[col_contacts] = pd.to_numeric(
                    df[col_contacts],
                    errors="coerce"
                )

                corr_vent = df[
                    [col_contacts, "vent"]
                ].corr().iloc[0,1]

                corr_temp = df[
                    [col_contacts, "température"]
                ].corr().iloc[0,1]

                resume = pd.DataFrame({
                    "Indicateur":[
                        "Corrélation contacts / vent",
                        "Corrélation contacts / température",
                        "Fuseau horaire déclaré - fichier brut",
                        "Fuseau horaire déclaré - fichier météo",
                        "Fuseau horaire des données après traitement"
                    ],
                    "Valeur":[
                        corr_vent,
                        corr_temp,
                        self.tz_excel.currentText(),
                        self.tz_csv.currentText(),
                        "UTC+2"
                    ]
                })

            self.pb.setValue(90)

            # ----------------------------------
            # export
            # ----------------------------------
            if self.ref_year:
                out_year = self.ref_year
            else:
                try:
                    out_year = int(
                        df["date_heure"].dt.year.mode().iloc[0]
                    )
                except Exception:
                    out_year = dtmod.date.today().year

            park_part = self.park_name if self.park_name else "Albat"

            default_name = f"{park_part}_{out_year}_Correlations.xlsx"

            if self.default_save_dir and os.path.isdir(self.default_save_dir):
                default_name = os.path.join(
                    self.default_save_dir, default_name
                )

            out, _ = QFileDialog.getSaveFileName(
                self,
                "Enregistrer résultat",
                default_name,
                "Excel (*.xlsx)"
            )

            if not out:
                return

            if not out.lower().endswith(".xlsx"):
                out += ".xlsx"

            with pd.ExcelWriter(out) as writer:

                df.to_excel(
                    writer,
                    sheet_name="Données",
                    index=False
                )

                if not resume.empty:
                    resume.to_excel(
                        writer,
                        sheet_name="Résumé",
                        index=False
                    )

            wb_out = openpyxl.load_workbook(out)

            self._style_worksheet(wb_out["Données"])
            self._hide_redundant_columns(
                wb_out["Données"], ["Date", "Heure"]
            )

            ws_data = wb_out["Données"]
            info_col = ws_data.max_column + 2
            info_letter = get_column_letter(info_col)
            info_letter2 = get_column_letter(info_col + 1)

            ws_data[f"{info_letter}1"] = "Parc éolien"
            ws_data[f"{info_letter2}1"] = self.park_name or "-"
            ws_data[f"{info_letter}2"] = "Année"
            ws_data[f"{info_letter2}2"] = self.ref_year or "-"

            for cell_ref in [f"{info_letter}1", f"{info_letter}2"]:
                ws_data[cell_ref].font = Font(name="Arial", bold=True)

            if "Résumé" in wb_out.sheetnames:
                self._style_worksheet(wb_out["Résumé"])

            wb_out.save(out)

            self.pb.setValue(100)

            open_file(out)

            QMessageBox.information(
                self,
                "Succès",
                "Analyse terminée."
            )

        except Exception as e:

            QMessageBox.critical(
                self,
                "Erreur",
                str(e))


# ----------------------------------------------------------
if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    fen = BatCorrelations()
    fen.show()

    sys.exit(app.exec())